#!/usr/bin/env python
# Test script for Excel template generation

import os
os.chdir('c:\\xampp2\\htdocs\\GescoApp')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    
    print("✅ openpyxl imported successfully")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Courses'
    
    headers = ['course_title', 'section_name', 'level', 'class_name', 'professor_full_name']
    ws.append(headers)
    
    print("✅ Headers appended")
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    print("✅ Header formatting applied")
    
    # Example rows
    ws.append(['Mathématiques', 'Section A', '1ère', 'A', 'John Doe'])
    ws.append(['Histoire'])
    ws.append(['Français', 'Section B', '2nde', 'B'])
    
    print("✅ Example rows appended")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    print("✅ Column widths adjusted")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    print("✅ Excel file saved to BytesIO")
    print(f"✅ File size: {len(output.getvalue())} bytes")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
