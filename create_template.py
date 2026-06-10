#!/usr/bin/env python
# Create template Excel file manually

import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Courses'

headers = ['course_title', 'section_name', 'level', 'class_name', 'professor_full_name']
ws.append(headers)

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Example rows
ws.append(['Mathématiques', 'Section A', '1ère', 'A', 'John Doe'])
ws.append(['Histoire'])
ws.append(['Français', 'Section B', '2nde', 'B'])
ws.append(['Biologie', '', '', '', ''])
ws.append(['Chimie'])

# Auto-fit columns
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

wb.save('template_courses.xlsx')
print("✅ Template Excel created: template_courses.xlsx")
