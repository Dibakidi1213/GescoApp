from flask import Blueprint, render_template, request, jsonify, redirect, url_for, g, session, send_file
from flask_login import login_required, current_user, logout_user
from sqlalchemy import func
from models import db, Course, Student, Grade, Section, BulletinConfig, BulletinBranch, User, AttendanceRecord, Notification, School
from routes.attendance_utils import class_denomination, class_denomination_from_parts
from url_utils import decode_id_or_int
from datetime import datetime, date, timedelta
from io import BytesIO
import re
import unicodedata
import openpyxl

professor_bp = Blueprint('professor', __name__, template_folder='../templates')

PERIODS = ['1èP', '2èP', 'EXA1', '3èP', '4èP', 'EXA2', 'REPECHAGE']
PERIOD_FIELD_MAP = {
    '1èP': 'max_period_1',
    '2èP': 'max_period_2',
    'EXA1': 'max_exam_1',
    '3èP': 'max_period_3',
    '4èP': 'max_period_4',
    'EXA2': 'max_exam_2'
}

PERIOD_INCLUDE_MAP = {
    '1èP': 'include_period_1',
    '2èP': 'include_period_2',
    'EXA1': 'include_comp_1',
    '3èP': 'include_period_3',
    '4èP': 'include_period_4',
    'EXA2': 'include_comp_2'
}

ATTENDANCE_STATUSES = {'present', 'absent', 'malade'}
ATTENDANCE_STATUS_LABELS = {
    'present': 'Présent',
    'absent': 'Absent',
    'malade': 'Malade',
}
ATTENDANCE_COURSE_TITLE_PREFIX = 'Présence de classe'


def _is_discipline_user():
    return hasattr(current_user, 'is_discipline') and current_user.is_discipline()


def _is_professor_user():
    return current_user.is_professor()


def _course_scope_query(school_id):
    query = Course.query.filter(
        Course.school_id == school_id,
        Course.section_id.isnot(None)
    )
    if _is_professor_user():
        query = query.filter(Course.professor_id == current_user.id)
    return query


def _attendance_scope_label(section):
    if not section:
        return None
    return class_denomination(section)


def _attendance_course_title(period):
    return f'{ATTENDANCE_COURSE_TITLE_PREFIX} - {period}'


def _format_attendance_date(value):
    if not value:
        return ''
    return value.strftime('%d/%m/%Y')


def _attendance_period_label(start_day, end_day):
    return f"du {_format_attendance_date(start_day)} au {_format_attendance_date(end_day)}"


def _parse_required_attendance_date(raw_value):
    raw = str(raw_value or '').strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None


def _attendance_print_days(start_day, end_day, include_saturday=True):
    days = []
    current_day = start_day
    while current_day <= end_day:
        weekday = current_day.weekday()
        is_saturday = weekday == 5
        is_sunday = weekday == 6
        if not is_sunday and (include_saturday or not is_saturday):
            days.append(current_day)
        current_day += timedelta(days=1)
    return days


def _attendance_gender_label(gender):
    normalized = str(gender or '').strip().upper()
    if normalized == 'M':
        return 'Masculin'
    if normalized == 'F':
        return 'Féminin'
    return ''


def _get_attendance_course_for_section(section, school_id, period, create=False):
    if not section:
        return None

    course = Course.query.filter_by(
        school_id=school_id,
        section_id=section.id,
        title=_attendance_course_title(period),
    ).first()
    if course or not create:
        return course

    course = Course(
        school_id=school_id,
        section_id=section.id,
        professor_id=current_user.id,
        title=_attendance_course_title(period),
    )
    db.session.add(course)
    db.session.flush()
    return course


def _resolve_attendance_section(school_id, section_name=None, level=None, class_name=None):
    if not section_name or not level or not class_name:
        return None

    sections = Section.query.filter_by(school_id=school_id).all()
    return next((
        section for section in sections
        if _section_matches_hierarchy(section, section_name, level, class_name)
    ), None)


def _resolve_attendance_course(school_id, course_id=None, section_name=None, level=None, class_name=None):
    if course_id:
        course_query = Course.query.filter_by(id=course_id, school_id=school_id)
        course = course_query.first()
        if course and course.section_id:
            return course

    section = _resolve_attendance_section(school_id, section_name, level, class_name)
    if not section:
        return None

    return Course.query.filter_by(
        school_id=school_id,
        section_id=section.id
    ).order_by(Course.id.asc()).first()


def _build_attendance_payload(course=None, section=None, attendance_day=None, year=None, period=None):
    if course is None and section is None:
        return None

    if course is not None:
        section = getattr(course, 'section', None)
        students = Student.query.filter_by(
            school_id=course.school_id,
            section_id=course.section_id,
            academic_year=year
        ).order_by(Student.last_name, Student.first_name).all()

        records_query = AttendanceRecord.query.filter_by(
            school_id=course.school_id,
            course_id=course.id,
            attendance_date=attendance_day,
            academic_year=year
        )
        if period:
            records_query = records_query.filter_by(period=period)
    else:
        students = Student.query.filter_by(
            school_id=section.school_id,
            section_id=section.id,
            academic_year=year
        ).order_by(Student.last_name, Student.first_name).all()

        records_query = AttendanceRecord.query.filter_by(
            school_id=section.school_id,
            section_id=section.id,
            attendance_date=attendance_day,
            academic_year=year
        )
        if period:
            records_query = records_query.filter_by(period=period)

    records = records_query.all()
    record_by_student = {record.student_id: record for record in records}

    payload_students = []
    for student in students:
        record = record_by_student.get(student.id)
        status = record.status if record and record.status in ATTENDANCE_STATUSES else 'present'
        payload_students.append({
            'student_id': student.id,
            'student_name': student.full_name(),
            'gender': student.gender,
            'status': status,
            'status_label': ATTENDANCE_STATUS_LABELS.get(status, 'Présent'),
        })

    return {
        'course_id': course.id if course else None,
        'course_title': course.title if course else None,
        'attendance_date': attendance_day.isoformat(),
        'period': period,
        'academic_year': year,
        'section_name': section.name if section else None,
        'level': section.level if section else None,
        'class_name': section.class_name if section else None,
        'class_label': _attendance_scope_label(section),
        'students': payload_students,
    }


def _normalize_text(value):
    compact = re.sub(r'\s+', ' ', str(value or '').strip())
    normalized = unicodedata.normalize('NFKD', compact)
    stripped = ''.join(char for char in normalized if not unicodedata.combining(char))
    return stripped.lower()


def _normalize_lookup_value(value):
    if value is None:
        return None
    return str(value).strip()


def _lower_trimmed(value):
    normalized = _normalize_lookup_value(value)
    return normalized.lower() if normalized is not None else None


def _coerce_branch_value(branch, field_name, fallback_field=None, default_value=0):
    value = getattr(branch, field_name, None)
    if value is None or (isinstance(value, str) and str(value).strip() == ''):
        if fallback_field:
            value = getattr(branch, fallback_field, None)
    if value is None or (isinstance(value, str) and str(value).strip() == ''):
        value = default_value
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default_value)


def _branch_period_limits(branch):
    return {
        '1èP': _coerce_branch_value(branch, 'max_period_1', default_value=10),
        '2èP': _coerce_branch_value(branch, 'max_period_2', default_value=10),
        'EXA1': _coerce_branch_value(branch, 'max_exam_1', default_value=20),
        '3èP': _coerce_branch_value(branch, 'max_period_3', default_value=10),
        '4èP': _coerce_branch_value(branch, 'max_period_4', default_value=10),
        'EXA2': _coerce_branch_value(branch, 'max_exam_2', default_value=20),
        'REPECHAGE': 100.0,
    }


def _get_repechage_eligible_student_ids(school_id, year, student_ids, course=None):
    if not student_ids:
        return set()
    from models import DeliberationResult
    results = DeliberationResult.query.filter(
        DeliberationResult.school_id == school_id,
        DeliberationResult.academic_year == year,
        DeliberationResult.period == 'ANNEE',
        DeliberationResult.student_id.in_(student_ids),
        DeliberationResult.decision == 'PASSAGE APRES REPECHAGE'
    ).all()
    if course and course.title:
        target = course.title.strip()
        return {
            result.student_id for result in results
            if any(note.strip() == target for note in (result.notes or '').split(';') if note.strip())
        }
    return {result.student_id for result in results}


def _parse_attendance_date(raw_value):
    raw = str(raw_value or '').strip()
    if not raw:
        return date.today()
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None


def _selection_from_course(course):
    if not course or not course.section:
        return {
            'section_name': None,
            'level': None,
            'class_name': None
        }

    return {
        'section_name': course.section.name,
        'level': course.section.level,
        'class_name': course.section.class_name
    }


def _section_matches_hierarchy(section, section_name=None, level=None, class_name=None):
    if not section:
        return False
    if section_name and _normalize_text(section.name) != _normalize_text(section_name):
        return False
    if level and _normalize_text(section.level) != _normalize_text(level):
        return False
    if class_name and _normalize_text(section.class_name) != _normalize_text(class_name):
        return False
    return True


def _visible_sections_for_current_user(school_id):
    """Retourne les sections visibles pour l'utilisateur connecte.

    Pour les utilisateurs 'discipline' : toutes les sections de l'ecole.
    Pour les professeurs : sections liees a leurs cours (via professor_id).
    Filtre les classes accessibles au professeur selon ses cours attribués.
    """
    if _is_discipline_user():
        return Section.query.filter_by(school_id=school_id).order_by(
            Section.name, Section.level, Section.class_name
        ).all()

    # Filtre sur les cours non-attendance (gere les encodages corrompus avec LIKE multiple)
    attendance_prefix = ATTENDANCE_COURSE_TITLE_PREFIX  # 'Présence de classe'
    query = Section.query.join(Course, Course.section_id == Section.id).filter(
        Section.school_id == school_id,
        Course.school_id == school_id,
        Course.section_id.isnot(None),
        ~Course.title.like(f'{attendance_prefix}%'),
        ~Course.title.like('Pr_sence de classe%'),
        ~Course.title.like('Presence de classe%'),
    )
    if _is_professor_user():
        query = query.filter(Course.professor_id == current_user.id)

    sections = query.distinct().order_by(Section.name, Section.level, Section.class_name).all()

    return sections


def _matching_visible_sections(school_id, section_name=None, level=None, class_name=None):
    return [
        section for section in _visible_sections_for_current_user(school_id)
        if _section_matches_hierarchy(section, section_name, level, class_name)
    ]


def _unique_sorted_values(values):
    seen = set()
    unique_values = []
    for value in values:
        normalized_value = _normalize_text(value)
        if not normalized_value or normalized_value in seen:
            continue
        seen.add(normalized_value)
        unique_values.append(str(value).strip())
    return sorted(unique_values, key=_normalize_text)


def _get_course_branch(course):
    if not course:
        return None

    section = getattr(course, 'section', None)
    if not section:
        return getattr(course, 'branch', None)

    config = BulletinConfig.query.filter_by(
        school_id=course.school_id,
        section_id=section.id,
        level=section.level
    ).order_by(BulletinConfig.updated_at.desc(), BulletinConfig.id.desc()).first()
    if not config:
        config = BulletinConfig.query.filter_by(
            school_id=course.school_id,
            section_id=section.id,
            level=section.level
        ).order_by(BulletinConfig.updated_at.desc(), BulletinConfig.id.desc()).first()
    if not config:
        return getattr(course, 'branch', None)

    branches = config.branches.order_by(BulletinBranch.order, BulletinBranch.id).all()
    if not branches:
        return getattr(course, 'branch', None)

    course_branch_id = getattr(course, 'branch_id', None)
    if course_branch_id is not None:
        matched_branch = next((branch for branch in branches if branch.id == course_branch_id), None)
        if matched_branch:
            return matched_branch

    normalized_title = _normalize_text(course.title)
    for branch in branches:
        if _normalize_text(branch.name) == normalized_title:
            return branch

    if len(branches) == 1:
        return branches[0]

    fallback_branch = getattr(course, 'branch', None)
    if fallback_branch and getattr(fallback_branch, 'config_id', None) == getattr(config, 'id', None):
        return fallback_branch

    return None


def _serialize_notification(notification):
    return {
        'id': notification.id,
        'title': notification.title,
        'message': notification.message,
        'type': notification.notification_type,
        'is_read': bool(notification.is_read),
        'created_at': notification.created_at.isoformat() if notification.created_at else None,
        'read_at': notification.read_at.isoformat() if notification.read_at else None,
        'url': notification.url
    }


def _create_notification(*, school_id, recipient_id, title, message, notification_type, actor_id=None, url=None):
    notification = Notification(
        school_id=school_id,
        recipient_id=recipient_id,
        actor_id=actor_id,
        notification_type=notification_type,
        title=title,
        message=message,
        url=url,
    )
    db.session.add(notification)
    return notification


@professor_bp.before_request
def restrict_professor():
    if not current_user.is_authenticated:
        return redirect(url_for('auth.login'))

    g.school_id = g.school.id if getattr(g, 'school', None) else current_user.school_id

    if current_user.is_professor() and current_user.must_change_password:
        redirect_args = {}
        if getattr(g, 'school_slug', None):
            redirect_args['school_slug'] = g.school_slug
        return redirect(url_for('auth.change_password', **redirect_args))

    if getattr(g, 'school', None) and not g.school.is_active and not current_user.is_super_admin():
        logout_user()
        return redirect(url_for('auth.login'))

    if getattr(g, 'school_slug', None) and not current_user.is_super_admin():
        if not current_user.can_access_school(g.school_id):
            return redirect(url_for('auth.redirect_by_role'))

    # Allow administrators and secretaries access to shared export endpoints
    if current_user.is_admin() or current_user.is_secretary():
        if request.endpoint == 'professor.export_fiche_cotes':
            return

    if not (_is_professor_user() or _is_discipline_user()):
        return redirect(url_for('auth.login'))

@professor_bp.route('/')
def dashboard(school_slug=None):
    school_id = g.school_id
    school_slug = school_slug or (current_user.school.slug if current_user.school and current_user.school.slug else None)
    can_manage_grades = _is_professor_user()
    can_manage_attendance = _is_discipline_user()
    can_view_notifications = can_manage_grades

    if can_manage_attendance and not can_manage_grades:
        if school_slug:
            return redirect(url_for('discipline.dashboard', school_slug=school_slug))
        return redirect(url_for('auth.redirect_by_role'))

    # Query parameters are used to restore the hierarchical selection in the UI.
    section_name = _normalize_lookup_value(request.args.get('section_name', type=str))
    if not section_name:
        section_name = _normalize_lookup_value(request.args.get('section', type=str))

    level = _normalize_lookup_value(request.args.get('level', type=str))
    class_name = _normalize_lookup_value(request.args.get('class_name', type=str))
    if not class_name:
        class_name = _normalize_lookup_value(request.args.get('class', type=str))

    section_id = decode_id_or_int(request.args.get('section_id'))
    course_id = decode_id_or_int(request.args.get('course_id'))

    course = None
    students = []
    grades_by_period = {}
    submitted_periods = set()
    branch_limits = None
    branch_enabled = None
    branch_name = None
    repechage_eligible_student_ids = set()
    flagged_student_ids = set()
    flagged_grade_pairs = set()
    selected_section_name = section_name
    selected_level = level
    selected_class_name = class_name
    selected_course_id = None

    if not selected_section_name and section_id:
        section = db.session.get(Section, section_id)
        if section and section.school_id == school_id:
            selected_section_name = section.name
            selected_level = section.level
            selected_class_name = section.class_name

    if can_manage_grades:
        default_course_query = Course.query.join(Section).filter(
            Course.school_id == school_id,
            Course.section_id.isnot(None),
            Course.professor_id == current_user.id
        )

        if selected_section_name:
            default_course_query = default_course_query.filter(
                func.lower(func.trim(Section.name)) == _lower_trimmed(selected_section_name)
            )
        if selected_level:
            default_course_query = default_course_query.filter(
                func.lower(func.trim(Section.level)) == _lower_trimmed(selected_level)
            )
        if selected_class_name:
            default_course_query = default_course_query.filter(
                func.lower(func.trim(Section.class_name)) == _lower_trimmed(selected_class_name)
            )

        if course_id:
            default_course_query = default_course_query.filter(Course.id == course_id)

        default_course = default_course_query.order_by(Section.name, Section.level, Section.class_name, Course.id).first()
        if default_course and default_course.section:
            selected_course_id = default_course.id
            selected_section_name = default_course.section.name
            selected_level = default_course.section.level
            selected_class_name = default_course.section.class_name
            if not course_id:
                course_id = selected_course_id

    if course_id:
        course_query = Course.query.filter_by(id=course_id, school_id=school_id)
        if can_manage_grades:
            course_query = course_query.filter_by(professor_id=current_user.id)
        course = course_query.first()
        if course:
            selected_course_id = course.id
            selection = _selection_from_course(course)
            selected_section_name = selection['section_name']
            selected_level = selection['level']
            selected_class_name = selection['class_name']

            if can_manage_grades:
                branch = _get_course_branch(course)
                if branch:
                    branch_limits = _branch_period_limits(branch)
                    branch_enabled = dict(zip(PERIODS, [
                        bool(branch.include_period_1),
                        bool(branch.include_period_2),
                        bool(branch.include_comp_1),
                        bool(branch.include_period_3),
                        bool(branch.include_period_4),
                        bool(branch.include_comp_2),
                        True,
                    ]))
                    branch_name = branch.name
                year = session.get('academic_year', '2025 - 2026')
                students = Student.query.filter_by(
                    school_id=school_id,
                    section_id=course.section_id,
                    academic_year=year
                ).order_by(Student.last_name, Student.first_name).all()
                all_grades = Grade.query.filter_by(school_id=school_id, course_id=course.id, academic_year=year).all()
                flagged_student_ids = {gr.student_id for gr in all_grades if gr.flagged}
                flagged_grade_pairs = {(gr.student_id, gr.period) for gr in all_grades if gr.flagged}
                for grade in all_grades:
                    if grade.student_id not in grades_by_period:
                        grades_by_period[grade.student_id] = {}
                    grades_by_period[grade.student_id][grade.period] = float(grade.value)

                # Calcul des périodes verrouillées (toutes les notes de cette période sont submitted)
                period_submitted_flags = {}  # period -> [True/False, ...]
                for grade in all_grades:
                    period_submitted_flags.setdefault(grade.period, []).append(bool(grade.submitted))
                submitted_periods = {
                    p for p, flags in period_submitted_flags.items() if flags and all(flags)
                }

                student_ids = [student.id for student in students]
                repechage_eligible_student_ids = _get_repechage_eligible_student_ids(school_id, year, student_ids, course=course)

                # Lorsque la période REPECHAGE est sélectionnée, ne garder que les élèves
                # admis au repêchage pour ce cours précis (si une délibération existe déjà).
                if request.args.get('period') == 'REPECHAGE':
                    has_repechage_results = bool(
                        _get_repechage_eligible_student_ids(school_id, year, student_ids)
                    )
                    if has_repechage_results:
                        repechage_for_course = _get_repechage_eligible_student_ids(
                            school_id, year, student_ids, course=course
                        )
                        students = [st for st in students if st.id in repechage_for_course]

    selected_class_label = class_denomination_from_parts(selected_section_name, selected_level, selected_class_name)

    # --- Preload all hierarchical data server-side so dropdowns fill immediately without AJAX ---
    attendance_prefix = ATTENDANCE_COURSE_TITLE_PREFIX
    all_visible_sections = _visible_sections_for_current_user(school_id)

    # Unique sorted section names
    seen_set = set()
    all_section_names = []
    for s in all_visible_sections:
        n = s.name.strip()
        if n.lower() not in seen_set:
            seen_set.add(n.lower())
            all_section_names.append(n)
    all_section_names.sort()

    # Unique levels for selected section
    all_levels = []
    if selected_section_name:
        seen_set = set()
        for s in all_visible_sections:
            if s.name.strip().lower() == selected_section_name.strip().lower():
                lv = s.level.strip()
                if lv.lower() not in seen_set:
                    seen_set.add(lv.lower())
                    all_levels.append(lv)
        all_levels.sort()

    # Unique classes for selected section + level
    all_classes = []
    if selected_section_name and selected_level:
        seen_set = set()
        for s in all_visible_sections:
            if (s.name.strip().lower() == selected_section_name.strip().lower()
                    and s.level.strip().lower() == selected_level.strip().lower()):
                cn = s.class_name.strip()
                if cn.lower() not in seen_set:
                    seen_set.add(cn.lower())
                    all_classes.append(cn)
        all_classes.sort()

    # Courses for selected section + level + class
    all_courses_for_class = []
    if selected_section_name and selected_level and selected_class_name:
        section_ids_for_class = [s.id for s in all_visible_sections
                                  if (s.name.strip().lower() == selected_section_name.strip().lower()
                                      and s.level.strip().lower() == selected_level.strip().lower()
                                      and s.class_name.strip().lower() == selected_class_name.strip().lower())]
        if section_ids_for_class:
            cq = Course.query.filter(
                Course.school_id == school_id,
                Course.section_id.in_(section_ids_for_class),
                ~Course.title.like(f'{attendance_prefix}%'),
                ~Course.title.like('Pr_sence de classe%'),
                ~Course.title.like('Presence de classe%'),
            )
            if can_manage_grades:
                cq = cq.filter(Course.professor_id == current_user.id)
            all_courses_for_class = [{'id': c.id, 'title': c.title} for c in cq.order_by(Course.title).all()]

    # Students list for preload
    all_students_for_class = [{
        'id': st.id,
        'full_name': st.full_name(),
        'last_name': st.last_name,
        'first_name': st.first_name,
    } for st in students]

    notifications = []
    unread_notifications_count = 0
    if can_view_notifications:
        notifications = Notification.query.filter_by(
            school_id=school_id,
            recipient_id=current_user.id
        ).order_by(Notification.is_read.asc(), Notification.created_at.desc(), Notification.id.desc()).limit(5).all()
        unread_notifications_count = Notification.query.filter_by(
            school_id=school_id,
            recipient_id=current_user.id,
            is_read=False
        ).count()

    school_obj = getattr(g, 'school', None) or (current_user.school if current_user.school else None)
    school_name_display = school_obj.name if school_obj else ''
    is_home_school = bool(school_obj and current_user.school_id == school_obj.id)
    professor_name = current_user.full_name if hasattr(current_user, 'full_name') else current_user.username
    dashboard_title = f'Bienvenue {professor_name} - Saisie des cotes' if can_manage_grades else 'Gestion des présences et de la conduite'

    accessible_schools = []
    if not current_user.is_super_admin():
        accessible_schools = current_user.accessible_schools()

    titulaire_section = None
    titulaire_url = None
    if current_user.titulaire_section_id:
        titulaire_section = db.session.get(Section, current_user.titulaire_section_id)
        if titulaire_section and current_user.school:
            titulaire_url = url_for('titulaire.dashboard', school_slug=current_user.school.slug)

    return render_template('professor/dashboard.html',
                           course=course,
                           students=students,
                           grades_by_period=grades_by_period,
                           submitted_periods=submitted_periods,
                           periods=PERIODS,
                           selected_section_name=selected_section_name,
                           selected_level=selected_level,
                           selected_class_name=selected_class_name,
                           selected_course_id=selected_course_id,
                           branch_limits=branch_limits,
                           branch_enabled=branch_enabled,
                           branch_name=branch_name,
                           repechage_eligible_student_ids=list(repechage_eligible_student_ids) if course else [],
                           flagged_student_ids=flagged_student_ids if course else set(),
                           flagged_grade_pairs=flagged_grade_pairs if course else set(),
                           notifications=notifications,
                           unread_notifications_count=unread_notifications_count,
                           can_manage_grades=can_manage_grades,
                           can_manage_attendance=can_manage_attendance,
                           can_view_notifications=can_view_notifications,
                           school_slug=school_slug,
                           school_name=school_name_display,
                           is_home_school=is_home_school,
                           accessible_schools=accessible_schools,
                           conduite_url=url_for('discipline.dashboard', school_slug=school_slug) if can_manage_attendance and school_slug else None,
                           selected_class_label=selected_class_label,
                           all_section_names=all_section_names,
                           all_levels=all_levels,
                           all_classes=all_classes,
                           all_courses_for_class=all_courses_for_class,
                           all_students_for_class=all_students_for_class,
                           titulaire_section=titulaire_section,
                           titulaire_url=titulaire_url,
                           dashboard_title=dashboard_title)

@professor_bp.route('/api/sections')
@login_required
def get_sections(school_slug=None):
    """Get all section names available to the current role."""
    school_id = g.school_id
    sections = _visible_sections_for_current_user(school_id)
    return jsonify(_unique_sorted_values(section.name for section in sections))

@professor_bp.route('/api/levels')
@professor_bp.route('/api/levels/<section_name>')
@login_required
def get_levels(section_name=None, school_slug=None):
    """Get all levels available for a section name."""
    school_id = g.school_id
    section_name = _normalize_lookup_value(
        section_name
        or request.args.get('section_name', type=str)
        or request.args.get('section', type=str)
    )
    if not section_name:
        return jsonify([])

    sections = _matching_visible_sections(school_id, section_name=section_name)
    return jsonify(_unique_sorted_values(section.level for section in sections))

@professor_bp.route('/api/classes')
@professor_bp.route('/api/classes/<section_name>/<level>')
@login_required
def get_classes(section_name=None, level=None, school_slug=None):
    """Get all classes for a section name and level."""
    school_id = g.school_id
    section_name = _normalize_lookup_value(
        section_name
        or request.args.get('section_name', type=str)
        or request.args.get('section', type=str)
    )
    level = _normalize_lookup_value(level or request.args.get('level', type=str))
    if not section_name or not level:
        return jsonify([])

    sections = _matching_visible_sections(school_id, section_name=section_name, level=level)
    return jsonify(_unique_sorted_values(section.class_name for section in sections))

@professor_bp.route('/api/courses')
@professor_bp.route('/api/courses/<section_name>/<level>/<class_name>')
@login_required
def get_courses(section_name=None, level=None, class_name=None, school_slug=None):
    """Get all courses for a section/level/class."""
    school_id = g.school_id
    section_name = _normalize_lookup_value(
        section_name
        or request.args.get('section_name', type=str)
        or request.args.get('section', type=str)
    )
    level = _normalize_lookup_value(level or request.args.get('level', type=str))
    class_name = _normalize_lookup_value(
        class_name
        or request.args.get('class_name', type=str)
        or request.args.get('class', type=str)
    )
    if not section_name or not level or not class_name:
        return jsonify([])

    sections = _matching_visible_sections(school_id, section_name, level, class_name)
    section_ids = [section.id for section in sections]
    if not section_ids:
        return jsonify([])

    attendance_prefix = ATTENDANCE_COURSE_TITLE_PREFIX
    query = Course.query.filter(
        Course.school_id == school_id,
        Course.section_id.in_(section_ids),
        ~Course.title.like(f'{attendance_prefix}%'),
        ~Course.title.like('Pr_sence de classe%'),
        ~Course.title.like('Presence de classe%'),
    )
    if _is_professor_user():
        query = query.filter(Course.professor_id == current_user.id)

    courses = query.order_by(Course.title).all()

    return jsonify([{
        'id': c.id,
        'title': c.title,
        'section_id': c.section_id,
        'section_name': c.section.name if c.section else None,
        'level': c.section.level if c.section else None,
        'class_name': c.section.class_name if c.section else None
    } for c in courses])


@professor_bp.route('/api/students')
@professor_bp.route('/api/students/<section_name>/<level>/<class_name>')
@login_required
def get_students(section_name=None, level=None, class_name=None, school_slug=None):
    """Get students for a section/level/class."""
    school_id = g.school_id
    section_name = _normalize_lookup_value(
        section_name
        or request.args.get('section_name', type=str)
        or request.args.get('section', type=str)
    )
    level = _normalize_lookup_value(level or request.args.get('level', type=str))
    class_name = _normalize_lookup_value(
        class_name
        or request.args.get('class_name', type=str)
        or request.args.get('class', type=str)
    )
    if not section_name or not level or not class_name:
        return jsonify([])

    sections = _matching_visible_sections(school_id, section_name, level, class_name)
    section_ids = [section.id for section in sections]
    if not section_ids:
        return jsonify([])

    query = Student.query.filter(
        Student.school_id == school_id,
        Student.section_id.in_(section_ids),
    )

    # Students are already scoped to sections where the professor has courses.
    # No extra Course JOIN needed (avoids duplicates and missed students).
    students = query.order_by(Student.last_name, Student.first_name).all()
    return jsonify([{
        'id': student.id,
        'full_name': student.full_name(),
        'gender': student.gender,
    } for student in students])


@professor_bp.route('/api/attendance', methods=['GET'])
@professor_bp.route('/api/attendance/<oid:course_id>', methods=['GET'])
@login_required
def get_attendance(course_id=None, school_slug=None):
    school_id = g.school_id
    if not _is_discipline_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    query_course_id = course_id or request.args.get('course_id', type=int)
    section_name = _normalize_lookup_value(request.args.get('section_name', type=str))
    level = _normalize_lookup_value(request.args.get('level', type=str))
    class_name = _normalize_lookup_value(request.args.get('class_name', type=str))
    period = request.args.get('period', type=str)
    attendance_day = _parse_attendance_date(request.args.get('date'))

    if not attendance_day:
        return jsonify({'error': 'Date de présence invalide.'}), 400
    if query_course_id and not period:
        return jsonify({'error': 'Période requise.'}), 400

    course = None
    section = None
    if query_course_id:
        course = _resolve_attendance_course(
            school_id,
            course_id=query_course_id,
            section_name=section_name,
            level=level,
            class_name=class_name
        )
        if not course:
            return jsonify({'error': 'Aucun cours associé.'}), 403
        if not course.section_id:
            return jsonify({'error': 'Ce cours n\'est pas lié à une classe.'}), 400
    else:
        section = _resolve_attendance_section(school_id, section_name, level, class_name)
        if not section:
            return jsonify({'error': 'Aucune classe associée.'}), 404
        if not period:
            return jsonify({'error': 'Période requise.'}), 400

    year = session.get('academic_year', '2025 - 2026')
    payload = _build_attendance_payload(
        course=course,
        section=section,
        attendance_day=attendance_day,
        year=year,
        period=period,
    )
    return jsonify(payload or {'error': 'Impossible de charger la présence.'})


@professor_bp.route('/attendance/print', methods=['GET'])
@login_required
def print_attendance_period(school_slug=None):
    if not _is_discipline_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    query_course_id = request.args.get('course_id', type=int)
    section_name = _normalize_lookup_value(request.args.get('section_name', type=str))
    level = _normalize_lookup_value(request.args.get('level', type=str))
    class_name = _normalize_lookup_value(request.args.get('class_name', type=str))
    start_day = _parse_required_attendance_date(request.args.get('start_date'))
    end_day = _parse_required_attendance_date(request.args.get('end_date'))
    include_saturday_raw = str(request.args.get('include_saturday', '1')).strip().lower()
    include_saturday = include_saturday_raw not in {'0', 'false', 'no', 'non', 'off'}

    if not start_day or not end_day:
        return jsonify({'error': 'Période de présence invalide.'}), 400
    if end_day < start_day:
        return jsonify({'error': 'La date de fin doit être supérieure ou égale à la date de début.'}), 400

    course = None
    section = None
    if query_course_id:
        course = _resolve_attendance_course(
            school_id,
            course_id=query_course_id,
            section_name=section_name,
            level=level,
            class_name=class_name
        )
        if not course or not course.section_id:
            return jsonify({'error': 'Aucune classe associée.'}), 404
        section = course.section
    else:
        section = _resolve_attendance_section(school_id, section_name, level, class_name)
        if not section:
            return jsonify({'error': 'Aucune classe associée.'}), 404

    period_label = str(request.args.get('period') or '').strip()
    if not period_label:
        period_label = _attendance_period_label(start_day, end_day)

    year = session.get('academic_year', '2025 - 2026')
    students = Student.query.filter_by(
        school_id=school_id,
        section_id=section.id,
        academic_year=year
    ).order_by(Student.last_name, Student.first_name).all()

    print_days = _attendance_print_days(start_day, end_day, include_saturday=include_saturday)
    records_by_day_student = {}
    student_ids = [student.id for student in students]

    attendance_course = course or _get_attendance_course_for_section(section, school_id, period_label, create=False)
    if attendance_course and print_days and student_ids:
        records = AttendanceRecord.query.filter(
            AttendanceRecord.school_id == school_id,
            AttendanceRecord.course_id == attendance_course.id,
            AttendanceRecord.attendance_date.in_(print_days),
            AttendanceRecord.period == period_label,
            AttendanceRecord.academic_year == year,
            AttendanceRecord.student_id.in_(student_ids)
        ).all()
        records_by_day_student = {
            (record.attendance_date, record.student_id): record
            for record in records
        }

    days_payload = []
    for print_day in print_days:
        rows = []
        for student in students:
            record = records_by_day_student.get((print_day, student.id))
            status = record.status if record and record.status in ATTENDANCE_STATUSES else ''
            rows.append({
                'student_name': student.full_name(),
                'gender_label': _attendance_gender_label(student.gender),
                'status': status,
                'status_label': ATTENDANCE_STATUS_LABELS.get(status, '') if status else '',
            })
        days_payload.append({
            'date': print_day,
            'date_label': _format_attendance_date(print_day),
            'rows': rows,
        })

    return render_template(
        'professor/attendance_print.html',
        school=getattr(g, 'school', None) or current_user.school,
        class_label=_attendance_scope_label(section),
        period_label=period_label,
        start_date_label=_format_attendance_date(start_day),
        end_date_label=_format_attendance_date(end_day),
        academic_year=year,
        days=days_payload,
        students_count=len(students),
        include_saturday=include_saturday,
    )


@professor_bp.route('/api/attendance/save', methods=['POST'])
@login_required
def save_attendance(school_slug=None):
    data = request.get_json() or {}
    school_id = g.school_id
    course_id = data.get('course_id')
    entries = data.get('entries') or []
    attendance_day = _parse_attendance_date(data.get('attendance_date'))
    section_name = _normalize_lookup_value(data.get('section_name'))
    level = _normalize_lookup_value(data.get('level'))
    class_name = _normalize_lookup_value(data.get('class_name'))
    period = str(_normalize_lookup_value(data.get('period')) or '').strip()

    try:
        course_id = int(course_id) if course_id not in (None, '') else None
    except (TypeError, ValueError):
        course_id = None

    if not attendance_day:
        return jsonify({'error': 'Date de présence invalide.'}), 400
    if not isinstance(entries, list):
        return jsonify({'error': 'entries doit être une liste.'}), 400
    if not period:
        return jsonify({'error': 'Période requise.'}), 400
    if not course_id and not (section_name and level and class_name):
        return jsonify({'error': 'section_name, level et class_name sont requis.'}), 400

    if not _is_discipline_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    section = None
    course = None
    if course_id:
        course = _resolve_attendance_course(
            school_id,
            course_id=course_id,
            section_name=section_name,
            level=level,
            class_name=class_name
        )
        if not course:
            return jsonify({'error': 'Aucun cours associé.'}), 403
        if not course.section_id:
            return jsonify({'error': 'Ce cours n\'est pas lié à une classe.'}), 400
        section = course.section
    else:
        section = _resolve_attendance_section(school_id, section_name, level, class_name)
        if not section:
            return jsonify({'error': 'Aucune classe associée.'}), 404
        course = _get_attendance_course_for_section(section, school_id, period, create=True)
        if not course:
            return jsonify({'error': 'Impossible de préparer la classe pour la présence.'}), 500

    year = session.get('academic_year', '2025 - 2026')
    students = Student.query.filter_by(
        school_id=school_id,
        section_id=section.id if section else course.section_id,
        academic_year=year
    ).all()
    valid_student_ids = {student.id for student in students}
    if not valid_student_ids:
        return jsonify({'error': 'Aucun élève trouvé pour cette classe.'}), 404

    existing_records = AttendanceRecord.query.filter(
        AttendanceRecord.school_id == school_id,
        AttendanceRecord.course_id == course.id,
        AttendanceRecord.attendance_date == attendance_day,
        AttendanceRecord.period == period,
        AttendanceRecord.academic_year == year,
        AttendanceRecord.student_id.in_(valid_student_ids)
    ).all()
    existing_by_student = {record.student_id: record for record in existing_records}

    created = 0
    updated = 0
    skipped = 0

    for item in entries:
        if not isinstance(item, dict):
            skipped += 1
            continue

        student_id_raw = item.get('student_id')
        try:
            student_id = int(student_id_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue

        if student_id not in valid_student_ids:
            skipped += 1
            continue

        status = str(item.get('status', 'present')).strip().lower()
        if status not in ATTENDANCE_STATUSES:
            skipped += 1
            continue

        record = existing_by_student.get(student_id)
        if record:
            record.status = status
            record.professor_id = current_user.id
            updated += 1
            continue

        db.session.add(AttendanceRecord(
            school_id=school_id,
            section_id=section.id if section else course.section_id,
            course_id=course.id,
            student_id=student_id,
            professor_id=current_user.id,
            attendance_date=attendance_day,
            period=period,
            status=status,
            academic_year=year
        ))
        created += 1

    db.session.commit()
    scope_label = _attendance_scope_label(course.section) if course.section else course.title
    return jsonify({
        'success': True,
        'attendance_date': attendance_day.isoformat(),
        'class_label': scope_label,
        'created': created,
        'updated': updated,
        'skipped': skipped
    })

@professor_bp.route('/grade', methods=['POST'])
@login_required
def save_grade(school_slug=None):
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    data = request.get_json() or {}
    student_id = data.get('student_id')
    value = data.get('value')
    period = data.get('period', '1èP')
    course_id = data.get('course_id')
    school_id = g.school_id
    
    # Validate course belongs to professor
    course = Course.query.filter_by(id=course_id, school_id=school_id, professor_id=current_user.id).first()
    if not course:
        return jsonify({'error': 'Aucun cours associé.'}), 403

    # Validate student belongs to the course's section (no cross-class writes)
    student = Student.query.filter_by(
        id=student_id, school_id=school_id, section_id=course.section_id
    ).first()
    if not student:
        return jsonify({'error': 'Élève introuvable dans cette classe.'}), 403

    # Check if this grade is locked (submitted) by the secretary
    year = session.get('academic_year', '2025 - 2026')
    existing_grade = Grade.query.filter_by(school_id=school_id, student_id=student_id, 
                                          course_id=course.id, period=period, academic_year=year).first()
    if existing_grade and existing_grade.submitted and not current_user.is_secretary():
        return jsonify({'error': f'Les notes de {period} ont été verrouillées par le secrétariat. Seul le secrétaire peut les modifier.'}), 403

    try:
        value = float(value)
    except (TypeError, ValueError):
        return jsonify({'error': 'Valeur de note invalide.'}), 400

    if value < 0:
        return jsonify({'error': 'La note ne peut pas être inférieure à 0.'}), 400

    branch = _get_course_branch(course)
    if branch:
        if period != 'REPECHAGE' and period not in PERIOD_FIELD_MAP:
            return jsonify({'error': f'Période invalide : {period}'}), 400

        if period != 'REPECHAGE' and not getattr(branch, PERIOD_INCLUDE_MAP.get(period), True):
            return jsonify({'error': f'La période {period} n’est pas active pour ce cours.'}), 400

        max_allowed = _branch_period_limits(branch).get(period, 100 if period == 'REPECHAGE' else 20)
        if value > max_allowed:
            return jsonify({'error': f'La note maximale pour {period} est {max_allowed}.'}), 400

        if period == 'REPECHAGE' and student_id not in _get_repechage_eligible_student_ids(school_id, year, [student_id]):
            return jsonify({'error': 'L’élève n’est pas éligible au repêchage.'}), 403

    grade = Grade.query.filter_by(school_id=school_id, student_id=student_id, 
                                  course_id=course.id, period=period, academic_year=year).first()
    if grade is None:
        grade = Grade(school_id=school_id, student_id=student_id, course_id=course.id, 
                      period=period, value=value, submitted=False, academic_year=year)
        db.session.add(grade)
    else:
        grade.value = value
    db.session.commit()
    return jsonify({'success': True, 'student_id': student_id, 'period': period, 'value': float(value), 'submitted': grade.submitted})


def _save_grade_draft_entry(course, school_id, student_id, period, value, year):
    # Validate student belongs to the course's section (no cross-class writes)
    student = Student.query.filter_by(
        id=student_id, school_id=school_id, section_id=course.section_id
    ).first()
    if not student:
        return None, {'error': "Élève introuvable dans cette classe."}, 403

    existing_grade = Grade.query.filter_by(
        school_id=school_id,
        student_id=student_id,
        course_id=course.id,
        period=period,
        academic_year=year
    ).first()

    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return None, {'error': 'Valeur de note invalide.'}, 400

    if existing_grade and existing_grade.submitted and not current_user.is_secretary():
        if existing_grade.value is not None and float(existing_grade.value) == numeric_value:
            # Si la note est verrouillée mais n'a pas été modifiée, on l'ignore sans erreur
            return existing_grade, None, None
        return None, {'error': f'Les notes de {period} ont ete verrouillees par le secretariat. Seul le secretaire peut les modifier.'}, 403

    if numeric_value < 0:
        return None, {'error': 'La note ne peut pas etre inferieure a 0.'}, 400

    branch = _get_course_branch(course)
    if branch:
        if period != 'REPECHAGE' and period not in PERIOD_FIELD_MAP:
            return None, {'error': f'Periode invalide : {period}'}, 400

        if period != 'REPECHAGE' and not getattr(branch, PERIOD_INCLUDE_MAP.get(period), True):
            return None, {'error': f"La periode {period} n'est pas active pour ce cours."}, 400

        max_allowed = _branch_period_limits(branch).get(period, 100 if period == 'REPECHAGE' else 20)
        if numeric_value > max_allowed:
            return None, {'error': f'La note maximale pour {period} est {max_allowed}.'}, 400

        if period == 'REPECHAGE' and student_id not in _get_repechage_eligible_student_ids(school_id, year, [student_id]):
            return None, {'error': "L'élève n'est pas éligible au repêchage."}, 403

    if existing_grade is None:
        grade = Grade(
            school_id=school_id,
            student_id=student_id,
            course_id=course.id,
            period=period,
            value=numeric_value,
            submitted=False,
            academic_year=year
        )
        db.session.add(grade)
    else:
        grade = existing_grade
        grade.value = numeric_value

    return grade, None, None


@professor_bp.route('/api/save-grades-draft', methods=['POST'])
@login_required
def save_grades_draft(school_slug=None):
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    data = request.get_json() or {}
    course_id = data.get('course_id')
    entries = data.get('grades') or []
    school_id = g.school_id

    if not course_id:
        return jsonify({'error': 'course_id est requis.'}), 400
    if not isinstance(entries, list):
        return jsonify({'error': 'grades doit etre une liste.'}), 400

    course = Course.query.filter_by(id=course_id, school_id=school_id, professor_id=current_user.id).first()
    if not course:
        return jsonify({'error': 'Aucun cours associe.'}), 403

    year = session.get('academic_year', '2025 - 2026')
    saved_count = 0
    skipped_count = 0

    for item in entries:
        if not isinstance(item, dict):
            skipped_count += 1
            continue

        student_id = item.get('student_id')
        period = item.get('period', '1èP')
        value = item.get('value')

        if value in (None, ''):
            skipped_count += 1
            continue

        _grade, error_payload, error_status = _save_grade_draft_entry(
            course=course,
            school_id=school_id,
            student_id=student_id,
            period=period,
            value=value,
            year=year
        )
        if error_payload:
            db.session.rollback()
            return jsonify(error_payload), error_status

        saved_count += 1

    db.session.commit()
    return jsonify({
        'success': True,
        'saved_count': saved_count,
        'skipped_count': skipped_count,
        'message': f'{saved_count} cotes enregistrees en brouillon.'
    })


@professor_bp.route('/api/submit-period', methods=['POST'])
@login_required
def submit_period(school_slug=None):
    """Submit all grades of a period for a course and publish them to the bulletin."""
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    data = request.get_json() or {}
    course_id = data.get('course_id')
    period = data.get('period')
    school_id = g.school_id

    if not course_id or not period:
        return jsonify({'error': 'course_id et period sont requis.'}), 400

    course = Course.query.filter_by(id=course_id, school_id=school_id, professor_id=current_user.id).first()
    if not course:
        return jsonify({'error': 'Aucun cours associé.'}), 403

    if period not in PERIODS:
        return jsonify({'error': f'Période invalide : {period}'}), 400

    # Get all grades for this period and mark them as submitted
    year = session.get('academic_year', '2025 - 2026')
    grades = Grade.query.filter_by(
        school_id=school_id,
        course_id=course.id,
        period=period,
        academic_year=year
    ).all()
    if not grades:
        return jsonify({'error': f'Aucune note trouvée pour la période {period}.'}), 404

    for grade in grades:
        grade.submitted = True
        grade.submitted_at = datetime.now()
        grade.submitted_by = current_user.id

    professor_name = getattr(current_user, 'full_name', None) or getattr(current_user, 'username', None) or 'Le professeur'
    secretaries = User.query.filter_by(school_id=school_id, role='secretary').order_by(User.full_name).all()
    if not secretaries:
        secretaries = User.query.filter_by(school_id=school_id, role='school_admin').order_by(User.full_name).all()

    for secretary in secretaries:
        _create_notification(
            school_id=school_id,
            recipient_id=secretary.id,
            actor_id=current_user.id,
            notification_type='period_submitted_by_professor',
            title='Notes envoyées au bulletin',
            message=(
                f'{professor_name} a publié {len(grades)} note(s) de {period} '
                f'dans le bulletin pour le cours {course.title} ({year}).'
            ),
        )

    db.session.commit()
    return jsonify({
        'success': True,
        'message': f'{len(grades)} notes de {period} ont été envoyées au bulletin et verrouillées pour le cours {course.title}.',
        'course_id': course.id,
        'period': period,
        'count': len(grades)
    })


@professor_bp.route('/api/get-flagged-grades/<oid:course_id>', methods=['GET'])
@login_required
def get_flagged_grades(course_id, school_slug=None):
    """Return flagged (student_id, period) pairs for a course, so the professor
    sees in real time which grades were modified by the titulaire/secretary."""
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    course = Course.query.filter_by(id=course_id, school_id=school_id, professor_id=current_user.id).first()
    if not course:
        return jsonify({'error': 'Aucun cours associé.'}), 403

    year = session.get('academic_year', '2025 - 2026')
    flagged = Grade.query.filter_by(
        school_id=school_id,
        course_id=course.id,
        academic_year=year,
        flagged=True,
    ).all()

    return jsonify([
        {'student_id': gr.student_id, 'period': gr.period, 'value': float(gr.value) if gr.value is not None else None}
        for gr in flagged
    ])


@professor_bp.route('/api/get-period-status/<oid:course_id>', methods=['GET'])
@login_required
def get_period_status(course_id, school_slug=None):
    """Get submission status for each period of a course"""
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    course = Course.query.filter_by(id=course_id, school_id=school_id, professor_id=current_user.id).first()
    if not course:
        return jsonify({'error': 'Aucun cours associé.'}), 403

    requested_period = request.args.get('period', type=str)
    if requested_period and requested_period not in PERIODS:
        return jsonify({'error': f'Période invalide : {requested_period}'}), 400

    status = {}
    year = session.get('academic_year', '2025 - 2026')
    for period in PERIODS:
        submitted_count = Grade.query.filter_by(
            school_id=school_id,
            course_id=course.id,
            period=period,
            submitted=True,
            academic_year=year
        ).count()
        total_count = Grade.query.filter_by(
            school_id=school_id,
            course_id=course.id,
            period=period,
            academic_year=year
        ).count()
        
        status[period] = {
            'submitted': total_count > 0 and submitted_count == total_count,
            'submitted_count': submitted_count,
            'total_count': total_count,
            'percentage': (submitted_count / total_count * 100) if total_count > 0 else 0
        }

    if requested_period:
        return jsonify({requested_period: status[requested_period]})

    return jsonify(status)


@professor_bp.route('/api/export-fiche-cotes', methods=['GET'])
@login_required
def export_fiche_cotes(school_slug=None):
    """Export Excel sheets for blank (fiche vierge) and filled (fiche remplie) fiches.
    Query parameters:
        course_id: ID of the course (required for filled fiche)
        filled: 0 for blank template, 1 for populated data (default 0)
    """
    if not (_is_professor_user() or current_user.is_admin() or current_user.is_secretary()):
        return jsonify({'error': 'Accès refusé.'}), 403

    filled = request.args.get('filled', default='0')
    try:
        filled = int(filled)
    except ValueError:
        filled = 0
    course_id = decode_id_or_int(request.args.get('course_id'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fiche'

    headers = ['Nom', 'Prénom', 'Classe'] + PERIODS
    ws.append(headers)
    header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    if filled == 1 and course_id:
        school_id = g.school_id
        course = Course.query.filter_by(id=course_id, school_id=school_id, professor_id=current_user.id).first()
        if not course:
            return jsonify({'error': 'Cours introuvable ou non autorisé.'}), 404
        year = session.get('academic_year', '2025 - 2026')
        students = Student.query.filter_by(school_id=school_id, section_id=course.section_id, academic_year=year)
        students = students.order_by(Student.last_name, Student.first_name).all()
        grades = Grade.query.filter_by(school_id=school_id, course_id=course.id, academic_year=year).all()
        grades_by_student = {}
        for grade in grades:
            grades_by_student.setdefault(grade.student_id, {})[grade.period] = grade.value
        for s in students:
            row = [s.last_name, s.first_name, class_denomination(course.section)]
            for p in PERIODS:
                row.append(grades_by_student.get(s.id, {}).get(p, ''))
            ws.append(row)
    else:
        ws.append(['', '', ''] + ['' for _ in PERIODS])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"fiche_{'remplie' if filled == 1 else 'vierge'}.xlsx"
    return send_file(file_stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@professor_bp.route('/api/export-grades', methods=['GET'])
@login_required
def export_grades_model(school_slug=None):
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    is_template = request.path.endswith('/export-template') or request.args.get('template_only') in ('1', 'true', 'True')

    if not (_is_professor_user() or current_user.is_admin() or current_user.is_secretary()):
        return jsonify({'error': 'Accès refusé.'}), 403

    course_id = decode_id_or_int(request.args.get('course_id'))
    
    school_id = g.school_id
    course = _course_scope_query(school_id).filter(Course.id == course_id).first()
    if not course:
        return jsonify({'error': 'Cours introuvable.'}), 404

    year = session.get('academic_year', '2025 - 2026')
    students = Student.query.filter_by(school_id=school_id, section_id=course.section_id, academic_year=year)
    students = students.order_by(Student.last_name, Student.first_name).all()

    grades = Grade.query.filter_by(school_id=school_id, course_id=course.id, academic_year=year).all()
    grades_by_student = {}
    for grade in grades:
        grades_by_student.setdefault(grade.student_id, {})[grade.period] = grade.value

    branch = _get_course_branch(course)
    limits = _branch_period_limits(branch) if branch else {}
    
    max_1ep = limits.get('1èP', 30)
    max_2ep = limits.get('2èP', 30)
    max_exa1 = limits.get('EXA1', 60)
    max_tot1 = max_1ep + max_2ep + max_exa1

    max_3ep = limits.get('3èP', 30)
    max_4ep = limits.get('4èP', 30)
    max_exa2 = limits.get('EXA2', 60)
    max_tot2 = max_3ep + max_4ep + max_exa2

    max_tot_gen = max_tot1 + max_tot2

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fiche des points'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 8

    # Headers and static text
    school = School.query.get(g.school_id) if g.school_id else None
    school_name = (school.bulletin_school_name or school.name) if school else ''
    other_denomination = (school.other_denomination or '').strip() if school else ''
    school_commune_province = ' / '.join(
        part for part in [
            (school.commune or '').strip() if school else '',
            (school.province or '').strip() if school else '',
        ] if part
    )

    header_lines = [school_name]
    commune_province_row = None
    if other_denomination:
        header_lines.extend([ln.strip() for ln in other_denomination.splitlines() if ln.strip()])
    if school_commune_province:
        header_lines.append(school_commune_province)
        commune_province_row = len(header_lines)

    offset = max(0, len(header_lines) - 4)
    for i, line in enumerate(header_lines, start=1):
        ws.cell(row=i, column=1, value=line)
        font_kwargs = {'bold': (i == 1)}
        if commune_province_row and i == commune_province_row:
            font_kwargs['underline'] = 'single'
        ws.cell(row=i, column=1).font = Font(**font_kwargs)


    ws['H1'] = f"Année scolaire {year}"
    ws.merge_cells('H1:L1')
    ws['H1'].font = Font(bold=True)
    ws['H1'].alignment = Alignment(horizontal='right')

    class_name = class_denomination(course.section) if course.section else ''
    ws['H2'] = class_name
    ws.merge_cells('H2:L3')
    ws['H2'].font = Font(bold=True, size=12)
    ws['H2'].alignment = Alignment(horizontal='center', vertical='center')
    
    thick_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )
    for row in ws['H2:L3']:
        for cell in row:
            cell.border = thick_border

    title_row = 5 + offset
    ws.cell(row=title_row, column=3, value="FICHE DES POINTS")
    ws.merge_cells(start_row=title_row, start_column=3, end_row=title_row + 1, end_column=10)
    ws.cell(row=title_row, column=3).font = Font(bold=True, size=16)
    ws.cell(row=title_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=title_row, max_row=title_row + 1, min_col=3, max_col=10):
        for cell in row:
            cell.border = thick_border

    branch_row = 8 + offset
    ws.cell(row=branch_row, column=1, value=f"Branche : {course.title}")
    ws.merge_cells(start_row=branch_row, start_column=1, end_row=branch_row, end_column=5)
    ws.cell(row=branch_row, column=1).font = Font(bold=True)
    
    professor_name = course.professor.full_name if course.professor else ""
    ws.cell(row=branch_row, column=8, value=f"Professeur : {professor_name}")
    ws.merge_cells(start_row=branch_row, start_column=8, end_row=branch_row, end_column=12)
    ws.cell(row=branch_row, column=8).font = Font(bold=True)
    ws.cell(row=branch_row, column=8).alignment = Alignment(horizontal='right')

    nb_row = 10 + offset
    ws.cell(row=nb_row, column=3, value="N.B : Les ratures et les surcharges ne sont pas tolérées")
    ws.merge_cells(start_row=nb_row, start_column=3, end_row=nb_row, end_column=10)
    ws.cell(row=nb_row, column=3).font = Font(bold=True, italic=True)
    ws.cell(row=nb_row, column=3).alignment = Alignment(horizontal='center')

    # Table headers
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)

    table_row = 12 + offset
    ws.cell(row=table_row, column=1, value="N°")
    ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row + 1, end_column=1)
    ws.cell(row=table_row, column=2, value="NOMS")
    ws.merge_cells(start_row=table_row, start_column=2, end_row=table_row + 1, end_column=2)
    ws.cell(row=table_row, column=3, value="POSTNOMS")
    ws.merge_cells(start_row=table_row, start_column=3, end_row=table_row + 1, end_column=3)
    
    ws.cell(row=table_row, column=4, value="1ère SEMESTRE")
    ws.merge_cells(start_row=table_row, start_column=4, end_row=table_row, end_column=7)
    ws.cell(row=table_row, column=8, value="2ème SEMESTRE")
    ws.merge_cells(start_row=table_row, start_column=8, end_row=table_row, end_column=11)
    
    ws.cell(row=table_row, column=12, value="TOTAL\nGENE")
    ws.merge_cells(start_row=table_row, start_column=12, end_row=table_row + 1, end_column=12)

    headers_13 = ['1èP', '2èP', 'Comp', 'TOT', '3èP', '4èP', 'Comp', 'TOT']
    for idx, h in enumerate(headers_13, start=4):
        ws.cell(row=table_row + 1, column=idx, value=h)

    for row in ws.iter_rows(min_row=table_row, max_row=table_row + 1, min_col=1, max_col=12):
        for cell in row:
            cell.border = thin_border
            cell.font = header_font
            cell.alignment = center_aligned

    max_row = table_row + 2
    ws.cell(row=max_row, column=1, value="")
    ws.merge_cells(start_row=max_row, start_column=1, end_row=max_row, end_column=2)
    ws.cell(row=max_row, column=3, value="MAXIMA")
    ws.cell(row=max_row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=max_row, column=3).font = header_font
    
    maxima_vals = [max_1ep, max_2ep, max_exa1, max_tot1, max_3ep, max_4ep, max_exa2, max_tot2, max_tot_gen]
    for idx, max_val in enumerate(maxima_vals, start=4):
        ws.cell(row=max_row, column=idx, value=max_val)

    for row in ws.iter_rows(min_row=max_row, max_row=max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = thin_border
            if cell.column >= 3:
                cell.alignment = center_aligned
                cell.font = header_font

    for i, s in enumerate(students, start=max_row + 1):
        ws.cell(row=i, column=1, value=i - 14)
        ws.cell(row=i, column=2, value=s.last_name)
        ws.cell(row=i, column=3, value=s.first_name)
        
        g_1ep = grades_by_student.get(s.id, {}).get('1èP', '') if not is_template else ''
        g_2ep = grades_by_student.get(s.id, {}).get('2èP', '') if not is_template else ''
        g_exa1 = grades_by_student.get(s.id, {}).get('EXA1', '') if not is_template else ''
        
        tot1 = ''
        if not is_template and (g_1ep != '' or g_2ep != '' or g_exa1 != ''):
            try:
                tot1 = sum([float(x) for x in [g_1ep, g_2ep, g_exa1] if str(x).strip() != ''])
                if tot1.is_integer(): tot1 = int(tot1)
            except:
                pass
            
        g_3ep = grades_by_student.get(s.id, {}).get('3èP', '') if not is_template else ''
        g_4ep = grades_by_student.get(s.id, {}).get('4èP', '') if not is_template else ''
        g_exa2 = grades_by_student.get(s.id, {}).get('EXA2', '') if not is_template else ''
        
        tot2 = ''
        if not is_template and (g_3ep != '' or g_4ep != '' or g_exa2 != ''):
            try:
                tot2 = sum([float(x) for x in [g_3ep, g_4ep, g_exa2] if str(x).strip() != ''])
                if tot2.is_integer(): tot2 = int(tot2)
            except:
                pass
            
        tot_gen = ''
        if not is_template and (tot1 != '' or tot2 != ''):
            try:
                tot_gen = sum([float(x) for x in [tot1, tot2] if str(x).strip() != ''])
                if tot_gen.is_integer(): tot_gen = int(tot_gen)
            except:
                pass

        vals = [g_1ep, g_2ep, g_exa1, tot1, g_3ep, g_4ep, g_exa2, tot2, tot_gen]
        
        for idx, val in enumerate(vals, start=4):
            c = ws.cell(row=i, column=idx, value=val)

        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        for col_idx in range(1, 13):
            c = ws.cell(row=i, column=col_idx)
            c.border = thin_border
            if col_idx >= 4:
                c.alignment = Alignment(horizontal='center')
                if is_template:
                    c.fill = gray_fill

    # --- Signature : Le Professeur / LA DIRECTION ---
    last_student_row = max_row + len(students)
    sig_line_row = last_student_row + 3

    ws.merge_cells(start_row=sig_line_row, start_column=2, end_row=sig_line_row, end_column=4)
    ws.cell(row=sig_line_row, column=2, value='_' * 24)
    ws.merge_cells(start_row=sig_line_row, start_column=9, end_row=sig_line_row, end_column=11)
    ws.cell(row=sig_line_row, column=9, value='_' * 24)

    ws.cell(row=sig_line_row + 1, column=2, value="Le Professeur")
    ws.cell(row=sig_line_row + 1, column=9, value="LA DIRECTION")
    for cell in (ws.cell(row=sig_line_row + 1, column=2), ws.cell(row=sig_line_row + 1, column=9)):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    prefix = "modele_vierge" if is_template else "fiche_des_points"
    filename = f"{prefix}_{class_name.strip().replace(' ', '_')}.xlsx"
    return send_file(file_stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@professor_bp.route('/api/import-grades', methods=['POST'])
@login_required
def import_grades(school_slug=None):
    if not (_is_professor_user() or current_user.is_admin() or current_user.is_secretary()):
        return jsonify({'error': 'Accès refusé.'}), 403

    course_id = request.form.get('course_id', type=int)
    period = request.form.get('period')
    
    if not period or (period not in ['1èP', '2èP', 'EXA1', '3èP', '4èP', 'EXA2'] and period != 'all'):
        return jsonify({'error': 'Période invalide.'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni.'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Format de fichier non supporté. Veuillez utiliser Excel (.xlsx).'}), 400

    school_id = g.school_id
    course = _course_scope_query(school_id).filter(Course.id == course_id).first()
    if not course:
        return jsonify({'error': 'Cours introuvable.'}), 404

    year = session.get('academic_year', '2025 - 2026')
    
    period_col_map = {
        '1èP': 4,
        '2èP': 5,
        'EXA1': 6,
        '3èP': 8,
        '4èP': 9,
        'EXA2': 10
    }
    
    periods_to_import = ['1èP', '2èP', 'EXA1', '3èP', '4èP', 'EXA2'] if period == 'all' else [period]

    branch = _get_course_branch(course)
    limits = _branch_period_limits(branch) if branch else {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        imported_count = 0
        errors = []
        
        for row_idx in range(15, ws.max_row + 1):
            last_name = ws.cell(row=row_idx, column=2).value
            first_name = ws.cell(row=row_idx, column=3).value
            
            if not last_name:
                continue 
                
            last_name = str(last_name).strip()
            first_name = str(first_name).strip() if first_name else ''
            
            student = Student.query.filter(
                Student.school_id == school_id,
                Student.section_id == course.section_id,
                Student.academic_year == year,
                db.func.lower(Student.last_name) == last_name.lower()
            )
            if first_name:
                student = student.filter(db.func.lower(Student.first_name) == first_name.lower())
            student = student.first()
                
            if not student:
                errors.append(f"Élève introuvable: {last_name} {first_name}")
                continue
                
            for p in periods_to_import:
                target_col = period_col_map.get(p)
                if not target_col:
                    continue

                grade_val = ws.cell(row=row_idx, column=target_col).value
                max_allowed = limits.get(p, 20)

                if grade_val is not None and str(grade_val).strip() != '':
                    try:
                        val = float(grade_val)
                        if val < 0 or val > max_allowed:
                            errors.append(f"Note invalide pour {last_name} en {p} (max {max_allowed})")
                            continue
                            
                        grade = Grade.query.filter_by(
                            school_id=school_id, 
                            student_id=student.id, 
                            course_id=course.id, 
                            period=p, 
                            academic_year=year
                        ).first()
                        
                        if grade and grade.submitted and not current_user.is_secretary():
                            errors.append(f"Notes verrouillées pour {last_name} en {p}")
                            continue
                            
                        if grade:
                            grade.value = val
                        else:
                            grade = Grade(
                                school_id=school_id,
                                student_id=student.id,
                                course_id=course.id,
                                period=p,
                                value=val,
                                academic_year=year,
                                submitted=False
                            )
                            db.session.add(grade)
                        imported_count += 1
                    except ValueError:
                        errors.append(f"Note non numérique pour {last_name} en {p}")
                    
        db.session.commit()
        
        error_msg = ""
        if errors:
            error_msg = " Quelques erreurs: " + ", ".join(errors[:3]) + ("..." if len(errors) > 3 else "")
            
        return jsonify({
            'success': True,
            'imported': imported_count,
            'error': error_msg if errors else None
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



@professor_bp.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications(school_slug=None):
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    limit = request.args.get('limit', default=5, type=int) or 5

    notifications = Notification.query.filter_by(
        school_id=school_id,
        recipient_id=current_user.id
    ).order_by(Notification.is_read.asc(), Notification.created_at.desc(), Notification.id.desc()).limit(limit).all()

    unread_count = Notification.query.filter_by(
        school_id=school_id,
        recipient_id=current_user.id,
        is_read=False
    ).count()

    return jsonify({
        'notifications': [_serialize_notification(notification) for notification in notifications],
        'unread_count': unread_count
    })


@professor_bp.route('/api/notifications/<oid:notification_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notification_id, school_slug=None):
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    notification = Notification.query.filter_by(
        id=notification_id,
        school_id=school_id,
        recipient_id=current_user.id
    ).first()

    if not notification:
        return jsonify({'error': 'Notification introuvable.'}), 404

    if not notification.is_read:
        notification.is_read = True
        notification.read_at = datetime.now()
        db.session.commit()

    return jsonify({
        'success': True,
        'notification': _serialize_notification(notification)
    })


@professor_bp.route('/api/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read(school_slug=None):
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    notifications = Notification.query.filter_by(
        school_id=school_id,
        recipient_id=current_user.id,
        is_read=False
    ).all()

    for notification in notifications:
        notification.is_read = True
        notification.read_at = datetime.now()

    db.session.commit()

    return jsonify({
        'success': True,
        'count': len(notifications)
    })


@professor_bp.route('/api/notifications/<oid:notification_id>/delete', methods=['POST'])
@login_required
def delete_notification(notification_id, school_slug=None):
    if not _is_professor_user():
        return jsonify({'error': 'Accès refusé.'}), 403

    school_id = g.school_id
    notification = Notification.query.filter_by(
        id=notification_id,
        school_id=school_id,
        recipient_id=current_user.id
    ).first()

    if not notification:
        return jsonify({'error': 'Notification introuvable.'}), 404

    db.session.delete(notification)
    db.session.commit()

    return jsonify({'success': True, 'id': notification_id})
