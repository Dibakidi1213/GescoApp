import os
from datetime import date, datetime

from flask import current_app, render_template, request, redirect, url_for, flash, session, jsonify
from flask_login import login_required, current_user

from models import AttendanceRecord, ActivityLog, Course, Grade, LoginHistory, Notification, Payment, School, SchoolHoliday, Section, Student, User, db
from routes.attendance_utils import (
    ATTENDANCE_STATUS_LABELS,
    ATTENDANCE_STATUS_OPTIONS,
    class_label,
    holiday_payload,
    is_working_day,
    month_bounds,
    month_key_for_day,
    parse_iso_date,
    parse_month,
    students_for_section_year,
)
from routes.admin.helpers import PERIOD_OPTIONS, get_school_id_for_admin_context, log_activity, require_super_admin, sum_payments_for_school
from routes.admin import admin_bp

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _render_placeholder(title, message='Cette fonctionnalité sera disponible prochainement.'):
    return render_template('admin/placeholder.html', title=title, message=message)


ATTENDANCE_STATUS_MAP = dict(ATTENDANCE_STATUS_OPTIONS)
ATTENDANCE_PERIOD_VALUES = {value for value, _label in PERIOD_OPTIONS}
ATTENDANCE_PLACEHOLDER_PREFIX = 'Présence de classe'


def _attendance_class_label(section):
    if not section:
        return ''
    return ' / '.join(part for part in [section.name, section.level, section.class_name] if part)


def _attendance_course_title(period):
    return f'{ATTENDANCE_PLACEHOLDER_PREFIX} - {period}'


def _get_attendance_course_for_section(section, school_id, period, create=False):
    if not section:
        return None

    course = Course.query.filter_by(
        school_id=school_id,
        section_id=section.id,
        title=_attendance_course_title(period),
    ).first()
    if course or not create:
        return course

    course = Course(
        school_id=school_id,
        section_id=section.id,
        title=_attendance_course_title(period),
        professor_id=current_user.id,
    )
    db.session.add(course)
    db.session.flush()
    return course


def _parse_attendance_date(raw_value):
    raw = str(raw_value or '').strip()
    if not raw:
        return date.today()
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None


def _current_academic_year():
    return session.get('academic_year', '2025 - 2026')


def _build_sections_catalog(sections):
    return [
        {
            'id': section.id,
            'name': section.name,
            'level': section.level,
            'class_name': section.class_name,
            'label': class_label(section),
        }
        for section in sections
    ]


def _day_holiday(school_id, attendance_day):
    if not attendance_day:
        return None
    return SchoolHoliday.query.filter_by(
        school_id=school_id,
        holiday_date=attendance_day,
    ).first()


def _record_map_for_day(school_id, section_ids, attendance_day, academic_year):
    if not section_ids:
        return {}

    month_key = month_key_for_day(attendance_day)
    records = AttendanceRecord.query.filter(
        AttendanceRecord.school_id == school_id,
        AttendanceRecord.section_id.in_(section_ids),
        AttendanceRecord.attendance_date == attendance_day,
        AttendanceRecord.academic_year == academic_year,
    ).order_by(AttendanceRecord.period.desc(), AttendanceRecord.id.desc()).all()

    record_map = {}
    for record in records:
        key = (record.section_id, record.student_id)
        if key not in record_map or record.period == month_key:
            record_map[key] = record
    return record_map


def _build_daily_stats(school_id, attendance_day, academic_year):
    sections = Section.query.filter_by(school_id=school_id).order_by(
        Section.name,
        Section.level,
        Section.class_name,
    ).all()
    section_ids = [section.id for section in sections]
    record_map = _record_map_for_day(school_id, section_ids, attendance_day, academic_year)
    holiday = _day_holiday(school_id, attendance_day)
    closed = not is_working_day(attendance_day, {holiday.holiday_date: holiday} if holiday else {})

    rows = []
    totals = {
        'students_count': 0,
        'entered_count': 0,
        'present_count': 0,
        'absent_count': 0,
        'malade_count': 0,
        'missing_count': 0,
    }

    for section in sections:
        students = students_for_section_year(school_id, section.id, academic_year)
        valid_student_ids = {student.id for student in students}
        counts = {
            'present_count': 0,
            'absent_count': 0,
            'malade_count': 0,
            'entered_count': 0,
        }
        for student_id in valid_student_ids:
            record = record_map.get((section.id, student_id))
            if not record:
                continue
            status = record.status
            if status == 'present':
                counts['present_count'] += 1
            elif status == 'absent':
                counts['absent_count'] += 1
            elif status == 'malade':
                counts['malade_count'] += 1
            counts['entered_count'] += 1

        students_count = len(students)
        missing_count = max(students_count - counts['entered_count'], 0)
        row = {
            'section_id': section.id,
            'class_label': class_label(section),
            'students_count': students_count,
            'entered_count': counts['entered_count'],
            'present_count': counts['present_count'],
            'absent_count': counts['absent_count'],
            'malade_count': counts['malade_count'],
            'missing_count': missing_count,
            'completion_rate': round((counts['entered_count'] / students_count * 100), 1) if students_count else 0,
        }
        rows.append(row)
        for key in totals:
            totals[key] += row[key]

    totals['completion_rate'] = round((totals['entered_count'] / totals['students_count'] * 100), 1) if totals['students_count'] else 0
    return {
        'attendance_date': attendance_day.isoformat() if attendance_day else '',
        'academic_year': academic_year,
        'closed': closed,
        'holiday': holiday_payload(holiday) if holiday else None,
        'rows': rows,
        'totals': totals,
    }


def _build_class_attendance_list(school_id, section, attendance_day, academic_year):
    students = students_for_section_year(school_id, section.id, academic_year)
    record_map = _record_map_for_day(school_id, [section.id], attendance_day, academic_year)
    rows = []
    counts = {
        'present_count': 0,
        'absent_count': 0,
        'malade_count': 0,
        'missing_count': 0,
    }

    for student in students:
        record = record_map.get((section.id, student.id))
        status = record.status if record else ''
        if status == 'present':
            counts['present_count'] += 1
        elif status == 'absent':
            counts['absent_count'] += 1
        elif status == 'malade':
            counts['malade_count'] += 1
        else:
            counts['missing_count'] += 1

        rows.append({
            'student_id': student.id,
            'student_name': student.full_name(),
            'gender': student.gender,
            'status': status,
            'status_label': ATTENDANCE_STATUS_LABELS.get(status, 'Non saisi') if status else 'Non saisi',
        })

    return {
        'section': {
            'id': section.id,
            'label': class_label(section),
            'name': section.name,
            'level': section.level,
            'class_name': section.class_name,
        },
        'attendance_date': attendance_day.isoformat() if attendance_day else '',
        'academic_year': academic_year,
        'rows': rows,
        'counts': counts,
        'students_count': len(students),
    }


@admin_bp.route('/professors', methods=['GET', 'POST'])
@login_required
def professors(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.dashboard', school_slug=school_slug))

    if request.method == 'POST':
        if request.files.get('import_file'):
            if not OPENPYXL_AVAILABLE:
                flash('Import Excel indisponible : installez openpyxl pour activer cette fonctionnalité.', 'warning')
                return redirect(url_for('admin.professors', school_slug=school_slug))

            workbook = openpyxl.load_workbook(request.files['import_file'])
            sheet = workbook.active
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [str(cell).strip().lower() if cell else '' for cell in (header_row or [])]
            full_name_index = next((i for i, name in enumerate(headers) if name in ('full_name', 'nom_complet', 'nom complet', 'nom')), None)
            username_index = next((i for i, name in enumerate(headers) if name in ('username', 'utilisateur', 'user_name')), None)
            email_index = next((i for i, name in enumerate(headers) if name in ('email', 'courriel')), None)
            password_index = next((i for i, name in enumerate(headers) if name in ('password', 'mot_de_passe', 'motdepasse')), None)

            if full_name_index is None:
                flash('Le fichier doit contenir une colonne <code>full_name</code> ou <code>nom_complet</code>.', 'danger')
                return redirect(url_for('admin.professors', school_slug=school_slug))

            created = 0
            skipped = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                full_name = str(row[full_name_index]).strip() if len(row) > full_name_index and row[full_name_index] else ''
                if not full_name:
                    skipped += 1
                    continue

                username = ''
                if username_index is not None and len(row) > username_index and row[username_index]:
                    username = str(row[username_index]).strip()
                if not username:
                    username = full_name

                if User.query.filter_by(username=username).first():
                    skipped += 1
                    continue

                email = str(row[email_index]).strip() if email_index is not None and len(row) > email_index and row[email_index] else None
                password = str(row[password_index]).strip() if password_index is not None and len(row) > password_index and row[password_index] else 'prof123'
                if not password:
                    password = 'prof123'

                professor = User(
                    school_id=school_id,
                    username=username,
                    full_name=full_name,
                    email=email,
                    role='professor',
                )
                professor.set_password(password)
                db.session.add(professor)
                created += 1

            db.session.commit()
            message = f'{created} professeur(s) importé(s) depuis Excel.'
            if skipped:
                message += f' {skipped} ligne(s) ignorée(s) en raison de données manquantes ou de doublons.'
            flash(message, 'success' if created else 'warning')
            return redirect(url_for('admin.professors', school_slug=school_slug))

        username = (request.form.get('username') or '').strip()
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        if not username or not password or not full_name:
            flash('Nom d\'utilisateur, mot de passe et nom complet sont obligatoires.', 'danger')
        elif User.query.filter_by(username=username).first():
            flash('Ce nom d\'utilisateur est déjà utilisé.', 'warning')
        else:
            professor = User(
                school_id=school_id,
                username=username,
                full_name=full_name,
                email=email,
                role='professor',
            )
            professor.set_password(password)
            db.session.add(professor)
            db.session.commit()
            flash('Professeur créé avec succès.', 'success')
        return redirect(url_for('admin.professors', school_slug=school_slug))

    professors_list = User.query.filter_by(school_id=school_id, role='professor').order_by(User.full_name).all()
    return render_template('admin/professors.html', professors=professors_list)


@admin_bp.route('/professors/<int:professor_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_professor(professor_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    professor = User.query.filter_by(id=professor_id, school_id=school_id, role='professor').first_or_404()

    if request.method == 'POST':
        professor.full_name = request.form.get('full_name') or professor.full_name
        professor.email = request.form.get('email')
        professor.username = request.form.get('username') or professor.username
        password = request.form.get('password')
        if password:
            professor.set_password(password)
        db.session.commit()
        flash('Professeur mis à jour.', 'success')
        return redirect(url_for('admin.professors', school_slug=school_slug))

    return render_template('admin/edit_professor.html', professor=professor)


@admin_bp.route('/professors/<int:professor_id>/delete', methods=['POST'])
@login_required
def delete_professor(professor_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    professor = User.query.filter_by(id=professor_id, school_id=school_id, role='professor').first_or_404()
    Course.query.filter_by(professor_id=professor.id).update({'professor_id': None})
    db.session.delete(professor)
    db.session.commit()
    flash('Professeur supprimé.', 'success')
    return redirect(url_for('admin.professors', school_slug=school_slug))


@admin_bp.route('/professors/<int:professor_id>/reset-password', methods=['POST'])
@login_required
def reset_professor_password(professor_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    professor = User.query.filter_by(id=professor_id, school_id=school_id, role='professor').first_or_404()
    professor.set_password('prof123')
    if hasattr(professor, 'must_change_password'):
        professor.must_change_password = True
    db.session.commit()
    flash('Mot de passe du professeur réinitialisé à prof123.', 'success')
    return redirect(url_for('admin.professors', school_slug=school_slug))


@admin_bp.route('/professors/reset-password', methods=['POST'])
@login_required
def reset_professor_password_form(school_slug=None):
    professor_id = request.form.get('professor_id')
    if not professor_id:
        flash('Veuillez sélectionner un professeur.', 'warning')
        return redirect(url_for('admin.school_reset', school_slug=school_slug))
    return reset_professor_password(int(professor_id), school_slug=school_slug)


@admin_bp.route('/school-reset')
@login_required
def school_reset(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.dashboard', school_slug=school_slug))

    sections = Section.query.filter_by(school_id=school_id).order_by(Section.name, Section.level, Section.class_name).all()
    section_names = sorted({section.name for section in sections})
    sections_catalog = [
        {
            'id': section.id,
            'name': section.name,
            'level': section.level,
            'class_name': section.class_name,
        }
        for section in sections
    ]
    professors = User.query.filter_by(school_id=school_id, role='professor').order_by(User.full_name).all()
    return render_template(
        'admin/school_reset.html',
        sections_catalog=sections_catalog,
        section_names=section_names,
        professors=professors,
    )


@admin_bp.route('/reset-course-assignments', methods=['POST'])
@login_required
def reset_course_assignments(school_slug=None):
    """Supprime tous les cours de l'école."""
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.school_reset', school_slug=school_slug))

    # Récupérer tous les IDs des cours de l'école
    courses = Course.query.filter_by(school_id=school_id).all()
    course_ids = [c.id for c in courses]

    if course_ids:
        # Supprimer les dépendances des cours (Grades et AttendanceRecords)
        Grade.query.filter(Grade.course_id.in_(course_ids)).delete(synchronize_session=False)
        AttendanceRecord.query.filter(AttendanceRecord.course_id.in_(course_ids)).delete(synchronize_session=False)
        # Supprimer les cours
        deleted = Course.query.filter(Course.id.in_(course_ids)).delete(synchronize_session=False)
        db.session.commit()
    else:
        deleted = 0

    flash(
        f'Attributions réinitialisées : {deleted} cours supprimés de la base de données.',
        'success' if deleted else 'info'
    )
    return redirect(url_for('admin.school_reset', school_slug=school_slug))


@admin_bp.route('/reset-section-course-assignments', methods=['POST'])
@login_required
def reset_section_course_assignments(school_slug=None):
    """Supprime tous les cours d'une section/niveau/classe spécifique."""
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.school_reset', school_slug=school_slug))

    section_name = (request.form.get('assign_section_name') or '').strip()
    level        = (request.form.get('assign_level') or '').strip()
    class_name   = (request.form.get('assign_class_name') or '').strip()

    if not section_name:
        flash('Veuillez sélectionner au moins une section.', 'warning')
        return redirect(url_for('admin.school_reset', school_slug=school_slug))

    # Construire le filtre sur les sections correspondantes
    section_query = Section.query.filter_by(school_id=school_id, name=section_name)
    if level:
        section_query = section_query.filter_by(level=level)
    if class_name:
        section_query = section_query.filter_by(class_name=class_name)
    matching_sections = section_query.all()

    if not matching_sections:
        flash('Aucune section trouvée pour ces critères.', 'warning')
        return redirect(url_for('admin.school_reset', school_slug=school_slug))

    section_ids = [s.id for s in matching_sections]
    
    # Récupérer tous les cours liés à ces sections
    courses = Course.query.filter(
        Course.school_id == school_id,
        Course.section_id.in_(section_ids)
    ).all()
    course_ids = [c.id for c in courses]

    if course_ids:
        # Supprimer les dépendances des cours (Grades et AttendanceRecords)
        Grade.query.filter(Grade.course_id.in_(course_ids)).delete(synchronize_session=False)
        AttendanceRecord.query.filter(AttendanceRecord.course_id.in_(course_ids)).delete(synchronize_session=False)
        # Supprimer les cours
        deleted = Course.query.filter(Course.id.in_(course_ids)).delete(synchronize_session=False)
        db.session.commit()
    else:
        deleted = 0

    label_parts = [section_name]
    if level:
        label_parts.append(f'Niveau {level}')
    if class_name:
        label_parts.append(f'Classe {class_name}')
    label = ' / '.join(label_parts)

    flash(
        f'Attributions réinitialisées pour {label} : {deleted} cours supprimés de la base de données.',
        'success' if deleted else 'info'
    )
    return redirect(url_for('admin.school_reset', school_slug=school_slug))


@admin_bp.route('/attendance')
@login_required
def attendance_management(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.dashboard', school_slug=school_slug))

    year = _current_academic_year()
    sections = Section.query.filter_by(school_id=school_id).order_by(
        Section.name,
        Section.level,
        Section.class_name,
    ).all()
    holidays = SchoolHoliday.query.filter_by(school_id=school_id).order_by(
        SchoolHoliday.holiday_date.desc(),
        SchoolHoliday.id.desc(),
    ).limit(200).all()
    return render_template(
        'admin/attendance.html',
        sections_catalog=_build_sections_catalog(sections),
        holidays=[holiday_payload(holiday) for holiday in holidays],
        today_iso=date.today().isoformat(),
        current_year=year,
    )


@admin_bp.route('/api/attendance/daily-stats', methods=['GET'])
@login_required
def attendance_daily_stats(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    attendance_day = parse_iso_date(request.args.get('date')) or date.today()
    stats = _build_daily_stats(school_id, attendance_day, _current_academic_year())
    return jsonify(stats)


@admin_bp.route('/api/attendance/class-list', methods=['GET'])
@login_required
def attendance_class_list(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    section_id = request.args.get('section_id', type=int)
    attendance_day = parse_iso_date(request.args.get('date')) or date.today()
    section = db.session.get(Section, section_id)
    if not section or section.school_id != school_id:
        return jsonify({'error': 'Classe introuvable.'}), 404
    return jsonify(_build_class_attendance_list(school_id, section, attendance_day, _current_academic_year()))


@admin_bp.route('/attendance/print-daily-stats', methods=['GET'])
@login_required
def print_attendance_daily_stats(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    attendance_day = parse_iso_date(request.args.get('date')) or date.today()
    stats = _build_daily_stats(school_id, attendance_day, _current_academic_year())
    return render_template(
        'admin/attendance_daily_print.html',
        school=current_user.school,
        stats=stats,
    )


@admin_bp.route('/attendance/print-class-list', methods=['GET'])
@login_required
def print_attendance_class_list(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    section_id = request.args.get('section_id', type=int)
    attendance_day = parse_iso_date(request.args.get('date')) or date.today()
    section = db.session.get(Section, section_id)
    if not section or section.school_id != school_id:
        flash('Classe introuvable.', 'danger')
        return redirect(url_for('admin.attendance_management', school_slug=school_slug))
    payload = _build_class_attendance_list(school_id, section, attendance_day, _current_academic_year())
    return render_template(
        'admin/attendance_class_print.html',
        school=current_user.school,
        payload=payload,
    )


@admin_bp.route('/api/attendance/holidays', methods=['POST'])
@login_required
def save_attendance_holiday(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    data = request.get_json(silent=True) or request.form
    holiday_date = parse_iso_date(data.get('holiday_date'))
    label = str(data.get('label') or '').strip()
    if not holiday_date:
        return jsonify({'error': 'Date de jour ferie invalide.'}), 400
    if not label:
        return jsonify({'error': 'Libelle obligatoire.'}), 400

    holiday = SchoolHoliday.query.filter_by(
        school_id=school_id,
        holiday_date=holiday_date,
    ).first()
    if holiday:
        holiday.label = label
        holiday.academic_year = _current_academic_year()
    else:
        holiday = SchoolHoliday(
            school_id=school_id,
            holiday_date=holiday_date,
            label=label,
            academic_year=_current_academic_year(),
        )
        db.session.add(holiday)
    db.session.commit()
    return jsonify({'success': True, 'holiday': holiday_payload(holiday)})


@admin_bp.route('/api/attendance/holidays/<int:holiday_id>', methods=['DELETE', 'POST'])
@login_required
def delete_attendance_holiday(holiday_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    holiday = SchoolHoliday.query.filter_by(id=holiday_id, school_id=school_id).first()
    if not holiday:
        return jsonify({'error': 'Jour ferie introuvable.'}), 404
    db.session.delete(holiday)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/attendance/class/<int:section_id>', methods=['GET'])
@login_required
def get_class_attendance(section_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        return jsonify({'error': 'Aucune école associée.'}), 400

    section = db.session.get(Section, section_id)
    if not section or section.school_id != school_id:
        return jsonify({'error': 'Classe non trouvée.'}), 404

    period = (request.args.get('period') or '').strip()
    if period not in ATTENDANCE_PERIOD_VALUES:
        return jsonify({'error': 'Période invalide.'}), 400

    attendance_day = _parse_attendance_date(request.args.get('date'))
    if not attendance_day:
        return jsonify({'error': 'Date de présence invalide.'}), 400

    year = session.get('academic_year', '2025 - 2026')
    students = Student.query.filter_by(
        school_id=school_id,
        section_id=section.id,
        academic_year=year,
    ).order_by(Student.last_name, Student.first_name).all()

    course = _get_attendance_course_for_section(section, school_id, period, create=False)
    records = []
    if course:
        records = AttendanceRecord.query.filter_by(
            school_id=school_id,
            course_id=course.id,
            attendance_date=attendance_day,
            period=period,
            academic_year=year,
        ).all()
    record_by_student = {record.student_id: record for record in records}

    payload_students = []
    for student in students:
        record = record_by_student.get(student.id)
        status = record.status if record and record.status in ATTENDANCE_STATUS_MAP else 'present'
        payload_students.append({
            'student_id': student.id,
            'full_name': student.full_name(),
            'gender': student.gender,
            'gender_label': 'Masculin' if student.gender == 'M' else 'Féminin' if student.gender == 'F' else '',
            'status': status,
            'status_label': ATTENDANCE_STATUS_MAP.get(status, 'Présent'),
        })

    return jsonify({
        'section': {
            'id': section.id,
            'name': section.name,
            'level': section.level,
            'class_name': section.class_name,
            'label': _attendance_class_label(section),
        },
        'attendance_date': attendance_day.isoformat(),
        'period': period,
        'academic_year': year,
        'course_id': course.id if course else None,
        'students': payload_students,
    })


@admin_bp.route('/api/attendance/class/<int:section_id>/save', methods=['POST'])
@login_required
def save_class_attendance(section_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        return jsonify({'error': 'Aucune école associée.'}), 400

    section = db.session.get(Section, section_id)
    if not section or section.school_id != school_id:
        return jsonify({'error': 'Classe non trouvée.'}), 404

    data = request.get_json() or {}
    attendance_day = _parse_attendance_date(data.get('attendance_date'))
    period = str(data.get('period') or '').strip()
    entries = data.get('entries') or []

    if not attendance_day:
        return jsonify({'error': 'Date de présence invalide.'}), 400
    if period not in ATTENDANCE_PERIOD_VALUES:
        return jsonify({'error': 'Période invalide.'}), 400
    if not isinstance(entries, list):
        return jsonify({'error': 'entries doit être une liste.'}), 400

    year = session.get('academic_year', '2025 - 2026')
    students = Student.query.filter_by(
        school_id=school_id,
        section_id=section.id,
        academic_year=year,
    ).all()
    valid_student_ids = {student.id for student in students}
    if not valid_student_ids:
        return jsonify({'error': 'Aucun élève trouvé pour cette classe.'}), 404

    course = _get_attendance_course_for_section(section, school_id, period, create=True)
    if not course:
        return jsonify({'error': 'Impossible de préparer la classe pour la présence.'}), 500

    existing_records = AttendanceRecord.query.filter_by(
        school_id=school_id,
        course_id=course.id,
        attendance_date=attendance_day,
        period=period,
        academic_year=year,
    ).all()
    existing_by_student = {record.student_id: record for record in existing_records}

    created = 0
    updated = 0
    skipped = 0

    for item in entries:
        if not isinstance(item, dict):
            skipped += 1
            continue

        try:
            student_id = int(item.get('student_id'))
        except (TypeError, ValueError):
            skipped += 1
            continue

        if student_id not in valid_student_ids:
            skipped += 1
            continue

        status = str(item.get('status', 'present')).strip().lower()
        if status not in ATTENDANCE_STATUS_MAP:
            skipped += 1
            continue

        record = existing_by_student.get(student_id)
        if record:
            record.status = status
            record.section_id = section.id
            record.professor_id = current_user.id
            updated += 1
            continue

        db.session.add(AttendanceRecord(
            school_id=school_id,
            section_id=section.id,
            course_id=course.id,
            student_id=student_id,
            professor_id=current_user.id,
            attendance_date=attendance_day,
            period=period,
            status=status,
            academic_year=year,
        ))
        created += 1

    db.session.commit()
    return jsonify({
        'success': True,
        'message': 'Présences enregistrées avec succès.',
        'attendance_date': attendance_day.isoformat(),
        'period': period,
        'class_label': _attendance_class_label(section),
        'created': created,
        'updated': updated,
        'skipped': skipped,
    })


@admin_bp.route('/exam-management')
@login_required
def exam_management(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    from models import Section
    sections = Section.query.filter_by(school_id=school_id).all() if school_id else []
    return render_template('admin/exam_management.html', sections=sections)


@admin_bp.route('/payments', methods=['GET', 'POST'])
@login_required
def payments(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.dashboard', school_slug=school_slug))

    if request.method == 'POST':
        student_id = request.form.get('student_id', type=int)
        amount = request.form.get('amount')
        payment_date_raw = request.form.get('payment_date')
        concept = request.form.get('concept')
        if not student_id or not amount:
            flash('Élève et montant obligatoires.', 'danger')
        else:
            payment = Payment(
                school_id=school_id,
                student_id=student_id,
                amount=amount,
                payment_date=datetime.strptime(payment_date_raw, '%Y-%m-%d').date() if payment_date_raw else date.today(),
                concept=concept,
            )
            db.session.add(payment)
            db.session.commit()
            flash('Paiement enregistré avec succès.', 'success')
        return redirect(url_for('admin.payments', school_slug=school_slug))

    payments_list = Payment.query.filter_by(school_id=school_id).order_by(Payment.payment_date.desc()).all()
    students = Student.query.filter_by(school_id=school_id).order_by(Student.last_name, Student.first_name).all()
    return render_template('admin/payments.html', payments=payments_list, students=students)


@admin_bp.route('/discipline')
@admin_bp.route('/cashier')
@login_required
def discipline_redirect(school_slug=None):
    target_slug = school_slug or (current_user.school.slug if current_user.school and current_user.school.slug else None)
    if current_user.is_discipline() or current_user.role == 'cashier':
        if target_slug:
            return redirect(url_for('discipline.dashboard', school_slug=target_slug))
        return redirect(url_for('auth.redirect_by_role'))
    return redirect(url_for('auth.redirect_by_role'))


@admin_bp.route('/audit-logs')
@login_required
@require_super_admin
def audit_logs(school_slug=None):
    school_id = get_school_id_for_admin_context()
    tab = request.args.get('tab', 'activity')

    now = datetime.utcnow()

    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    user_filter = request.args.get('user_id', type=int)
    action_filter = request.args.get('action_type', '')

    today = now.strftime('%Y-%m-%d')

    if tab == 'logins':
        query = LoginHistory.query
        if school_id:
            query = query.filter_by(school_id=school_id)
        if user_filter:
            query = query.filter_by(user_id=user_filter)
        if date_from:
            query = query.filter(LoginHistory.login_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            query = query.filter(LoginHistory.login_at < datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        query = query.order_by(LoginHistory.login_at.desc())
        page = request.args.get('page', 1, type=int)
        pagination = query.paginate(page=page, per_page=50, error_out=False)
        items = pagination.items
        total = pagination.total
        pages = pagination.pages
        current_page = page
    else:
        query = ActivityLog.query
        if school_id:
            query = query.filter_by(school_id=school_id)
        if user_filter:
            query = query.filter_by(user_id=user_filter)
        if action_filter:
            query = query.filter(ActivityLog.action_type == action_filter)
        if date_from:
            query = query.filter(ActivityLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        if date_to:
            query = query.filter(ActivityLog.created_at < datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        query = query.order_by(ActivityLog.created_at.desc())
        page = request.args.get('page', 1, type=int)
        pagination = query.paginate(page=page, per_page=50, error_out=False)
        items = pagination.items
        total = pagination.total
        pages = pagination.pages
        current_page = page

    action_types = [row[0] for row in db.session.query(ActivityLog.action_type).distinct().order_by(ActivityLog.action_type).all()]

    users_list = User.query.order_by(User.full_name).all()

    stats = _build_audit_stats(school_id)

    return render_template(
        'admin/audit_logs.html',
        tab=tab,
        items=items,
        total=total,
        pages=pages,
        current_page=current_page,
        date_from=date_from,
        date_to=date_to,
        user_filter=user_filter,
        action_filter=action_filter,
        action_types=action_types,
        users_list=users_list,
        today=today,
        stats=stats,
        school_id=school_id,
    )


def _build_audit_stats(school_id=None):
    base_activity = ActivityLog.query
    base_logins = LoginHistory.query
    if school_id:
        base_activity = base_activity.filter_by(school_id=school_id)
        base_logins = base_logins.filter_by(school_id=school_id)

    return {
        'total_activities': base_activity.count(),
        'total_logins': base_logins.count(),
        'successful_logins': base_logins.filter_by(success=True).count(),
        'failed_logins': base_logins.filter_by(success=False).count(),
        'unique_users_activity': db.session.query(ActivityLog.user_id).filter(
            ActivityLog.school_id == school_id if school_id else db.text('1=1')
        ).distinct().count() if school_id else base_activity.with_entities(ActivityLog.user_id).distinct().count(),
    }


@admin_bp.route('/backups')
@login_required
@require_super_admin
def backups(school_slug=None):
    from flask import current_app

    school_id = get_school_id_for_admin_context()

    db_type = db.engine.dialect.name

    db_size = 0
    db_mtime = None
    snapshots = []
    is_file_based = db_type == 'sqlite'

    if is_file_based:
        basedir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
        db_path = os.path.join(basedir, 'gescoapp.db')
        db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        db_mtime = datetime.fromtimestamp(os.path.getmtime(db_path)) if os.path.exists(db_path) else None

        backup_dir = os.path.join(basedir, 'backups')
        os.makedirs(backup_dir, exist_ok=True)

        if os.path.isdir(backup_dir):
            for f in sorted(os.listdir(backup_dir), reverse=True):
                fpath = os.path.join(backup_dir, f)
                if os.path.isfile(fpath) and f.endswith('.db'):
                    snapshots.append({
                        'filename': f,
                        'size': os.path.getsize(fpath),
                        'mtime': datetime.fromtimestamp(os.path.getmtime(fpath)),
                    })

    schools = School.query.order_by(School.name).all() if not school_id else School.query.filter_by(id=school_id).all()

    allow_restore = current_app.config.get('ALLOW_RESTORE_DOWNLOAD', False)

    return render_template(
        'admin/backups.html',
        db_exists=is_file_based,
        db_type=db_type,
        db_size=db_size,
        db_mtime=db_mtime,
        snapshots=snapshots,
        schools=schools,
        allow_restore=allow_restore,
    )


@admin_bp.route('/backups/create-snapshot', methods=['POST'])
@login_required
@require_super_admin
def create_backup_snapshot(school_slug=None):
    import os, shutil
    from datetime import datetime

    if db.engine.dialect.name != 'sqlite':
        flash('La création de snapshot est uniquement disponible avec SQLite.', 'warning')
        return redirect(url_for('admin.backups', school_slug=school_slug))

    basedir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    db_path = os.path.join(basedir, 'gescoapp.db')
    if not os.path.exists(db_path):
        flash('Base de données introuvable.', 'danger')
        return redirect(url_for('admin.backups', school_slug=school_slug))

    backup_dir = os.path.join(basedir, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    snapshot_name = f'snapshot_{timestamp}.db'
    snapshot_path = os.path.join(backup_dir, snapshot_name)

    try:
        shutil.copy2(db_path, snapshot_path)
        flash(f'Snapshot créé : {snapshot_name}', 'success')
    except Exception as e:
        flash(f'Erreur lors de la création du snapshot : {str(e)}', 'danger')

    return redirect(url_for('admin.backups', school_slug=school_slug))


@admin_bp.route('/backups/delete-snapshot/<filename>', methods=['POST'])
@login_required
@require_super_admin
def delete_backup_snapshot(filename, school_slug=None):
    import os
    if db.engine.dialect.name != 'sqlite':
        flash('La suppression de snapshot est uniquement disponible avec SQLite.', 'warning')
        return redirect(url_for('admin.backups', school_slug=school_slug))
    basedir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    snapshot_path = os.path.join(basedir, 'backups', os.path.basename(filename))
    if os.path.exists(snapshot_path):
        os.remove(snapshot_path)
        flash(f'Snapshot supprimé : {filename}', 'success')
    else:
        flash('Snapshot introuvable.', 'danger')
    return redirect(url_for('admin.backups', school_slug=school_slug))


@admin_bp.route('/communications')
@login_required
@require_super_admin
def communications(school_slug=None):
    school_id = get_school_id_for_admin_context()

    page = request.args.get('page', 1, type=int)
    school_filter = request.args.get('school_id', type=int)
    type_filter = request.args.get('type', '')

    query = Notification.query

    if school_id:
        query = query.filter_by(school_id=school_id)
    if school_filter:
        query = query.filter_by(school_id=school_filter)
    if type_filter:
        query = query.filter_by(notification_type=type_filter)
    else:
        query = query.filter(Notification.notification_type.in_(['global', 'system', 'general']))

    query = query.order_by(Notification.created_at.desc())
    pagination = query.paginate(page=page, per_page=50, error_out=False)

    schools = School.query.order_by(School.name).all()
    notification_types = [row[0] for row in db.session.query(Notification.notification_type).distinct().order_by(Notification.notification_type).all()]

    total_sent = Notification.query.count()
    total_unread = Notification.query.filter_by(is_read=False).count()

    return render_template(
        'admin/communications.html',
        notifications=pagination.items,
        total=total_sent,
        pages=pagination.pages,
        current_page=page,
        schools=schools,
        school_filter=school_filter,
        type_filter=type_filter,
        notification_types=notification_types,
        total_unread=total_unread,
        total_sent=total_sent,
    )


@admin_bp.route('/communications/send', methods=['POST'])
@login_required
@require_super_admin
def send_communication(school_slug=None):
    title = (request.form.get('title') or '').strip()
    message = (request.form.get('message') or '').strip()
    notification_type = (request.form.get('notification_type') or 'global').strip()
    target_school_id = request.form.get('school_id', type=int)
    target_role = request.form.get('target_role', '').strip()

    if not title or not message:
        flash('Le titre et le message sont obligatoires.', 'danger')
        return redirect(url_for('admin.communications', school_slug=school_slug))

    recipients = []

    if notification_type == 'direct' and target_school_id:
        school = School.query.get(target_school_id)
        if school:
            users_query = User.query.filter_by(school_id=school.id)
            if target_role:
                users_query = users_query.filter_by(role=target_role)
            recipients = [(school.id, u.id) for u in users_query.all()]
    else:
        schools_query = School.query.filter_by(is_active=True)
        if target_school_id:
            schools_query = schools_query.filter_by(id=target_school_id)
        for school in schools_query.all():
            users_query = User.query.filter_by(school_id=school.id)
            if target_role:
                users_query = users_query.filter_by(role=target_role)
            recipients.extend([(school.id, u.id) for u in users_query.all()])

    if not recipients:
        flash('Aucun destinataire trouvé pour cette cible.', 'warning')
        return redirect(url_for('admin.communications', school_slug=school_slug))

    for school_id_val, recipient_id in recipients:
        notification = Notification(
            school_id=school_id_val,
            recipient_id=recipient_id,
            actor_id=current_user.id,
            notification_type=notification_type,
            title=title,
            message=message,
        )
        db.session.add(notification)

    db.session.commit()

    log_activity(
        'communication_send',
        f'Communication envoyée : "{title}" à {len(recipients)} destinataire(s)',
        related_model='Notification',
        school_id=target_school_id or None,
    )

    flash(f'✅ Communication envoyée à {len(recipients)} destinataire(s).', 'success')
    return redirect(url_for('admin.communications', school_slug=school_slug))


@admin_bp.route('/global-reports')
@login_required
@require_super_admin
def global_reports(school_slug=None):
    return _render_placeholder('Rapports Globaux')
