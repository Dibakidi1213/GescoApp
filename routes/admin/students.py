import os
import re
import uuid
from datetime import datetime

from flask import render_template, request, redirect, url_for, flash, session
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from models import Grade, Payment, Section, Student, db
from routes.admin.helpers import get_school_id_for_admin_context
from routes.admin.services import group_students_tree
from routes.secretary import _get_section_by_hierarchy
from routes.admin import admin_bp

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _parse_student_line(line):
    raw = line.strip()
    if not raw:
        return None
    gender = None
    if ';' in raw:
        name_part, gender_part = raw.split(';', 1)
        raw_name = name_part.strip()
        gender = _normalize_gender(gender_part)
    elif '|' in raw:
        name_part, gender_part = raw.split('|', 1)
        raw_name = name_part.strip()
        gender = _normalize_gender(gender_part)
    else:
        raw_name = raw

    parts = raw_name.split()
    if not parts:
        return None
    last_name = parts[0]
    first_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
    return last_name, first_name, gender


def _normalize_gender(raw_gender):
    if raw_gender is None:
        return None
    value = str(raw_gender).strip().lower()
    if not value:
        return None
    if value.startswith('m'):
        return 'M'
    if value.startswith('f'):
        return 'F'
    if value in ('masculin', 'mâle', 'male', 'homme'):
        return 'M'
    if value in ('féminin', 'feminin', 'female', 'femme'):
        return 'F'
    return None


def _normalize_excel_header(label):
    if label is None:
        return None
    value = str(label).strip().lower()
    if not value:
        return None
    if value in ('nom', 'last_name', 'lastname', 'nom de famille', 'name'):
        return 'last_name'
    if value in ('postnom', 'first_name', 'firstname', 'prénom', 'prenom', 'given_name'):
        return 'first_name'
    if value in ('sexe', 'genre', 'gender'):
        return 'gender'
    if value in ('date de naissance', 'birth_date', 'birthday', 'dob'):
        return 'birth_date'
    return None


def _build_excel_header_map(header_row):
    return {
        _normalize_excel_header(header): idx
        for idx, header in enumerate(header_row)
        if _normalize_excel_header(header)
    }


def _parse_excel_student_row(row, header_map):
    values = [cell if cell is not None else '' for cell in row]

    if header_map:
        last_name = str(values[header_map['last_name']]).strip() if 'last_name' in header_map else str(values[0]).strip()
        first_name = ''
        if 'first_name' in header_map:
            first_name = str(values[header_map['first_name']]).strip()
        elif len(values) > 1:
            candidate = str(values[1]).strip()
            if _normalize_gender(candidate) is None:
                first_name = candidate
        gender = _normalize_gender(values[header_map['gender']]) if 'gender' in header_map else None
        if gender is None and len(values) > 2 and 'gender' not in header_map:
            gender = _normalize_gender(values[2])
        return last_name, first_name, gender

    last_name = str(values[0]).strip()
    first_name = ''
    gender = None
    if len(values) > 1:
        second = str(values[1]).strip()
        if _normalize_gender(second) is None:
            first_name = second
        else:
            gender = _normalize_gender(second)
    if len(values) > 2:
        gender = gender or _normalize_gender(values[2])
    return last_name, first_name, gender


def _resolve_target_section(school_id, section_name, level, class_name):
    return _get_section_by_hierarchy(school_id, section_name, level, class_name)


@admin_bp.route('/students', methods=['GET', 'POST'])
@login_required
def students(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    if not school_id:
        flash('Aucune école associée.', 'danger')
        return redirect(url_for('admin.dashboard', school_slug=school_slug))

    year = session.get('academic_year', '2025 - 2026')

    if request.method == 'POST':
        import_mode = request.form.get('student_import_mode')

        if import_mode == 'quick_list':
            section_name = request.form.get('quick_section_name')
            level = request.form.get('quick_level')
            class_name = request.form.get('quick_class_name')
            section = _resolve_target_section(school_id, section_name, level, class_name)
            if not section:
                flash('Section, niveau ou classe invalide.', 'danger')
                return redirect(url_for('admin.students', school_slug=school_slug))

            created = 0
            for line in (request.form.get('quick_students_list') or '').splitlines():
                parsed = _parse_student_line(line)
                if not parsed:
                    continue
                last_name, first_name, gender = parsed
                student = Student(
                    school_id=school_id,
                    section_id=section.id,
                    last_name=last_name,
                    first_name=first_name,
                    gender=gender or 'M',
                    academic_year=year,
                )
                db.session.add(student)
                created += 1
            db.session.commit()
            flash(f'{created} élève(s) importé(s) rapidement.', 'success')
            return redirect(url_for('admin.students', school_slug=school_slug))

        if request.files.get('import_file'):
            if not OPENPYXL_AVAILABLE:
                flash('Import Excel indisponible : installez openpyxl pour activer cette fonctionnalité.', 'warning')
                return redirect(url_for('admin.students', school_slug=school_slug))

            section_name = request.form.get('import_section_name')
            level = request.form.get('import_level')
            class_name = request.form.get('import_class_name')
            section = _resolve_target_section(school_id, section_name, level, class_name)
            if not section:
                flash('Section, niveau ou classe invalide pour l\'import.', 'danger')
                return redirect(url_for('admin.students', school_slug=school_slug))

            workbook = openpyxl.load_workbook(request.files['import_file'])
            sheet = workbook.active
            created = 0
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_map = _build_excel_header_map(first_row)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                last_name, first_name, gender = _parse_excel_student_row(row, header_map)
                if not last_name:
                    continue
                student = Student(
                    school_id=school_id,
                    section_id=section.id,
                    last_name=last_name,
                    first_name=first_name,
                    gender=gender or 'M',
                    academic_year=year,
                )
                db.session.add(student)
                created += 1
            db.session.commit()
            flash(f'{created} élève(s) importé(s) depuis Excel.', 'success')
            return redirect(url_for('admin.students', school_slug=school_slug))

        section_name = request.form.get('section_name')
        level = request.form.get('level')
        class_name = request.form.get('class_name')
        section = _resolve_target_section(school_id, section_name, level, class_name)
        if not section:
            flash('Veuillez sélectionner une section, un niveau et une classe valides.', 'danger')
            return redirect(url_for('admin.students', school_slug=school_slug))

        birth_date = request.form.get('birth_date')
        parsed_birth_date = None
        if birth_date:
            parsed_birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()

        student = Student(
            school_id=school_id,
            section_id=section.id,
            last_name=request.form.get('last_name'),
            first_name=request.form.get('first_name'),
            birth_date=parsed_birth_date,
            place_of_birth=request.form.get('place_of_birth'),
            gender=request.form.get('gender') or 'M',
            father_name=request.form.get('father_name'),
            mother_name=request.form.get('mother_name'),
            parent_phone=request.form.get('parent_phone'),
            student_id_number=request.form.get('student_id_number'),
            serial_number=request.form.get('serial_number'),
            address=request.form.get('address'),
            phone=request.form.get('phone'),
            email=request.form.get('email'),
            academic_year=year,
        )
        db.session.add(student)
        db.session.commit()
        flash('Élève enregistré avec succès.', 'success')
        return redirect(url_for('admin.students', school_slug=school_slug))

    grouped_students = group_students_tree(school_id, year)
    total_students = Student.query.filter_by(school_id=school_id, academic_year=year).count()
    sections = Section.query.filter_by(school_id=school_id).order_by(Section.name, Section.level, Section.class_name).all()
    section_names = sorted({section.name for section in sections})
    sections_catalog = [
        {
            'id': section.id,
            'name': section.name,
            'level': section.level,
            'class_name': section.class_name,
        }
        for section in sections
    ]

    return render_template(
        'admin/students.html',
        grouped_students=grouped_students,
        total_students=total_students,
        section_names=section_names,
        sections_catalog=sections_catalog,
        sections=sections,
    )


@admin_bp.route('/students/reset-class', methods=['POST'])
@login_required
def reset_student_class(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    year = session.get('academic_year', '2025 - 2026')
    section_name = request.form.get('reset_section_name_class')
    level = request.form.get('reset_level_class')
    class_name = request.form.get('reset_class_name')

    section = _resolve_target_section(school_id, section_name, level, class_name)
    if not section:
        flash('Section, niveau ou classe invalide pour la réinitialisation.', 'danger')
        return redirect(url_for('admin.students', school_slug=school_slug))

    deleted_count = Student.query.filter_by(
        school_id=school_id,
        section_id=section.id,
        academic_year=year,
    ).delete(synchronize_session=False)
    db.session.commit()

    flash(f'Réinitialisation effectuée : {deleted_count} élève(s) supprimé(s) de la classe {class_name} ({section_name} niveau {level}).', 'success')
    return redirect(url_for('admin.students', school_slug=school_slug))


@admin_bp.route('/students/reset-level', methods=['POST'])
@login_required
def reset_student_level(school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    year = session.get('academic_year', '2025 - 2026')
    section_name = request.form.get('reset_section_name_level')
    level = request.form.get('reset_level_level')

    sections = Section.query.filter_by(school_id=school_id, name=section_name, level=level).all()
    if not sections:
        flash('Section ou niveau invalide pour la réinitialisation.', 'danger')
        return redirect(url_for('admin.students', school_slug=school_slug))

    section_ids = [section.id for section in sections]
    deleted_count = Student.query.filter(
        Student.school_id == school_id,
        Student.section_id.in_(section_ids),
        Student.academic_year == year,
    ).delete(synchronize_session=False)
    db.session.commit()

    flash(f'Réinitialisation effectuée : {deleted_count} élève(s) supprimé(s) pour le niveau {level} de la section {section_name}.', 'success')
    return redirect(url_for('admin.students', school_slug=school_slug))


@admin_bp.route('/students/<int:student_id>/delete', methods=['POST'])
@login_required
def delete_student(student_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    student = Student.query.filter_by(id=student_id, school_id=school_id).first_or_404()
    Grade.query.filter_by(student_id=student.id).delete(synchronize_session=False)
    Payment.query.filter_by(student_id=student.id).delete(synchronize_session=False)
    db.session.delete(student)
    db.session.commit()
    flash('Élève supprimé avec succès.', 'success')
    return redirect(url_for('admin.students', school_slug=school_slug))


@admin_bp.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_student(student_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    student = Student.query.filter_by(id=student_id, school_id=school_id).first_or_404()
    sections = Section.query.filter_by(school_id=school_id).order_by(Section.name, Section.level, Section.class_name).all()

    if request.method == 'POST':
        student.last_name = request.form.get('last_name') or student.last_name
        student.first_name = request.form.get('first_name')
        student.place_of_birth = request.form.get('place_of_birth')
        student.gender = request.form.get('gender') or student.gender
        student.father_name = request.form.get('father_name')
        student.mother_name = request.form.get('mother_name')
        student.parent_phone = request.form.get('parent_phone')
        student.student_id_number = request.form.get('student_id_number')
        student.serial_number = request.form.get('serial_number')
        student.address = request.form.get('address')
        student.phone = request.form.get('phone')
        student.email = request.form.get('email')
        section_id = request.form.get('section_id', type=int)
        if section_id:
            student.section_id = section_id
        birth_date = request.form.get('birth_date')
        if birth_date:
            student.birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
        db.session.commit()
        flash('Informations de l\'élève mises à jour.', 'success')
        return redirect(url_for('admin.students', school_slug=school_slug))

    return render_template('admin/edit_student.html', student=student, sections=sections)


@admin_bp.route('/students/<int:student_id>')
@login_required
def student_detail(student_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    student = Student.query.filter_by(id=student_id, school_id=school_id).first_or_404()
    return render_template('admin/student_detail.html', student=student)


@admin_bp.route('/students/<int:student_id>/upload-photo', methods=['POST'])
@login_required
def upload_student_photo(student_id, school_slug=None):
    school_id = get_school_id_for_admin_context() or current_user.school_id
    student = Student.query.filter_by(id=student_id, school_id=school_id).first_or_404()
    photo = request.files.get('photo')
    if not photo or not photo.filename:
        flash('Veuillez sélectionner une photo.', 'warning')
        return redirect(url_for('admin.student_detail', student_id=student.id, school_slug=school_slug))

    filename = secure_filename(photo.filename)
    extension = os.path.splitext(filename)[1].lower()
    upload_dir = os.path.join('static', 'uploads', 'students')
    os.makedirs(upload_dir, exist_ok=True)
    stored_name = f"student_{student.id}_{uuid.uuid4().hex}{extension}"
    photo.save(os.path.join(upload_dir, stored_name))
    student.photo_url = f"/static/uploads/students/{stored_name}"
    db.session.commit()
    flash('Photo de l\'élève mise à jour.', 'success')
    return redirect(url_for('admin.student_detail', student_id=student.id, school_slug=school_slug))
